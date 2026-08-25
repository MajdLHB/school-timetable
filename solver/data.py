"""Load the school data from data/school.xlsx and sanity-check it.

This file knows nothing about solving. If check() is happy, the data is
internally consistent. That does NOT mean a timetable exists - only that the
numbers make sense.
"""
import json
import os
import sys
from dataclasses import dataclass, field

# Windows consoles default to cp1252, which cannot encode Arabic and raises
# UnicodeEncodeError mid-print. Force UTF-8 so real names are printable.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# The workbook has a grey hint row at row 2 which must be skipped.
FIRST_DATA_ROW = 2


@dataclass
class Config:
    days: list
    periods_per_day: int
    morning: list
    evening: list
    closed: dict
    time_limit: int
    weights: dict
    weeks_per_cycle: int = 1

    @property
    def slots(self):
        """Every (day, period) the school is open, in time order."""
        out = []
        for d in self.days:
            shut = set(self.closed.get(d, []))
            for p in range(1, self.periods_per_day + 1):
                if p not in shut:
                    out.append((d, p))
        return out

    def day_slots(self, day):
        return [(d, p) for (d, p) in self.slots if d == day]


@dataclass
class School:
    cfg: Config
    teachers: dict = field(default_factory=dict)
    classes: dict = field(default_factory=dict)
    rooms: dict = field(default_factory=dict)
    subjects: dict = field(default_factory=dict)
    curriculum: list = field(default_factory=list)
    unavailable: list = field(default_factory=list)
    locked: list = field(default_factory=list)
    # H14 (Majd's answers, 2026-08-25): option groups pool pupils from
    # several same-year classes; every pupil takes exactly one option, so
    # all option lessons of one pool run SIMULTANEOUSLY (a "band").
    options: list = field(default_factory=list)
    option_bands: list = field(default_factory=list)
    # which workbook this school was loaded from (stamped into the XML so
    # verify.py can refuse a timetable built from a different file)
    source_path: str = ""

    def room_type_for(self, cur_row):
        """Which kind of room this curriculum row needs."""
        if cur_row.get("room_type"):
            return cur_row["room_type"]
        subj = self.subjects.get(cur_row["subject_id"], {})
        return subj.get("room_type", "normal") or "normal"

    def rooms_of_type(self, t):
        return [r for r in self.rooms.values() if r["type"] == t]


def load_config(path=None):
    path = path or os.path.join(HERE, "config.json")
    with open(path, encoding="utf-8") as f:
        c = json.load(f)
    closed = {k: v for k, v in c.get("closed", {}).items() if not k.startswith("_")}
    return Config(
        days=c["days"],
        periods_per_day=c["periods_per_day"],
        morning=c["morning_periods"],
        evening=c["evening_periods"],
        closed=closed,
        time_limit=c.get("time_limit_seconds", 120),
        weights=c["weights"],
        weeks_per_cycle=c.get("weeks_per_cycle", 1),
    )


def _rows(ws):
    """Yield dicts from an openpyxl sheet, skipping the hint row and blanks."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    for r in rows[FIRST_DATA_ROW:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        rec = {}
        for h, v in zip(header, r):
            if h:
                rec[h] = "" if v is None else str(v).strip()
        if rec.get(header[0]):
            yield rec


def _int(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def parse_blocks(blocks, hours):
    """H9: '2+1+1' -> [2, 1, 1]. Blank -> all single hours.

    Returns (list, error_message). error_message is None when the value is
    readable; the CONSISTENCY of the list (sum vs hours, fitting the week)
    is judged in check(), where it can be reported in plain language.
    """
    t = str(blocks or "").strip()
    if not t:
        return ([1] * hours if hours > 0 else [], None)
    parts = [p for p in t.replace(" ", "").split("+") if p]
    try:
        out = [int(float(p)) for p in parts]
    except ValueError:
        return (None, "blocks '%s' is unreadable - write it like 2+1+1" % blocks)
    if not out or any(v <= 0 for v in out):
        return (None, "blocks '%s' must be positive numbers joined by +" % blocks)
    return (out, None)


def longest_open_run(cfg):
    """The longest run of NUMERICALLY consecutive open periods on any day.
    A block longer than this cannot be placed anywhere (the lunch break
    interrupts every day, so on the real config this is 4)."""
    best = 0
    for d in cfg.days:
        ps = sorted(p for (dd, p) in cfg.slots if dd == d)
        run = 0
        prev = None
        for p in ps:
            run = run + 1 if prev is not None and p == prev + 1 else 1
            prev = p
            best = max(best, run)
    return best


def load_school(xlsx=None, cfg=None):
    from openpyxl import load_workbook

    cfg = cfg or load_config()
    xlsx = xlsx or os.path.join(HERE, "data", "school.xlsx")
    if not os.path.exists(xlsx):
        sys.exit(
            "No data file at " + xlsx + "\n"
            "  python tools/make_workbook.py   -> blank workbook to fill in\n"
            "  python tools/make_demo.py       -> demo school, to try the tool"
        )
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    s = School(cfg=cfg)
    s.source_path = xlsx        # stamped into the XML, checked by verify.py

    for r in _rows(wb["Teachers"]):
        s.teachers[r["id"]] = dict(
            id=r["id"],
            name=r.get("name") or r["id"],
            short=r.get("short") or r["id"],
            subjects=[x for x in (r.get("subjects") or "").split(";") if x],
            hours=_int(r.get("hours")),
            day_off=(r.get("day_off") or "").strip(),
            # M-T6 / circular II.1: the pedagogical training day stays empty,
            # exactly like a day off. Blank = no training day.
            training_day=(r.get("training_day") or "").strip(),
            # S8: ministry default is hours spread over most days (II.2).
            # compact=yes keeps the packed week - the exception Majd grants to
            # teachers with long journeys. Blank = ministry default.
            compact=(r.get("compact") or "").strip().lower(),
            # S21 shared transport: the partner's teacher id. Filling one
            # side is enough - the pair is symmetric.
            travels_with=(r.get("travels_with") or "").strip(),
            notes=r.get("notes", ""),
        )
    for r in _rows(wb["Classes"]):
        s.classes[r["id"]] = dict(
            id=r["id"],
            name=r.get("name") or r["id"],
            grade=r.get("grade", ""),
            stream=(r.get("stream") or "").strip(),
            # yes = final year. Drives S13 (no Friday evening, local
            # preference) and S17 (free afternoon Mon-Thu, circular I.6).
            is_bac=(r.get("is_bac") or "").strip().lower(),
            cohort=(r.get("cohort") or "ALL").upper(),
            home_room=r.get("home_room", ""),
            size=_int(r.get("size")),
        )
    for r in _rows(wb["Rooms"]):
        s.rooms[r["id"]] = dict(
            id=r["id"],
            name=r.get("name") or r["id"],
            type=(r.get("type") or "normal").strip(),
            capacity=_int(r.get("capacity"), 999),
            # T45 room proximity: rooms in the same zone are close. Blank =
            # one big zone = the rule costs nothing (Majd fills them later).
            zone=(r.get("zone") or "").strip(),
        )
    for r in _rows(wb["Subjects"]):
        s.subjects[r["id"]] = dict(
            id=r["id"],
            name=r.get("name") or r["id"],
            short=r.get("short") or r["id"],
            difficulty=(r.get("difficulty") or "medium").strip(),
            room_type=(r.get("room_type") or "normal").strip(),
            # H15: last period this subject may occupy. Blank = no limit.
            # Older workbooks have no such column, hence the tolerant get().
            latest_period=_int(r.get("latest_period"), 0),
            # S16: soft version - prefer not to sit after this period.
            # Ministry: Maths before 16:00 (M-MA3), Physics not 17-18 (M-PH5).
            avoid_after=_int(r.get("avoid_after"), 0),
            # Circular I.2 note: the min-2h/session rules do not apply to PE
            # and optional subjects. yes = exempt from S15.
            minmax_exempt=(r.get("minmax_exempt") or "").strip().lower(),
            # H19, circular III.2 note on PE: always 24 hours between the two
            # sessions. yes = on consecutive days the later session must not
            # start earlier than the first one did.
            gap24=(r.get("gap24") or "").strip().lower(),
            # S18: subjects this one must not FOLLOW immediately (same class,
            # adjacent periods). The inspectorate: never Philosophy straight
            # after PE - so PHIL carries not_after=SPORT. ; separated ids.
            not_after=[v for v in (r.get("not_after") or "").replace(",", ";").split(";")
                       if v.strip()],
            # S4 / M-P6: the subject's nature (literary / scientific /
            # social). Two DIFFERENT subjects of the same nature back to
            # back are penalised; a double of one subject is fine.
            nature=(r.get("nature") or "").strip().lower(),
            # T37: subjects that must not share a DAY with this one (soft) -
            # the ministry's History/Geography rule. One side is enough.
            not_same_day=[v for v in (r.get("not_same_day") or "")
                          .replace(",", ";").split(";") if v.strip()],
        )
    # Majd 2026-08-25: "it did have tp and lesson for same subject day and
    # afternoon" - a subject's TP and its theory lesson must not share a
    # day. The <SID>_TP naming is ours, so the pair links automatically
    # (T37, weight not_same_day - editable in the Weights sheet).
    for sid, sub in s.subjects.items():
        if sid.endswith("_TP") and sid[:-3] in s.subjects:
            if sid[:-3] not in sub["not_same_day"]:
                sub["not_same_day"].append(sid[:-3])

    for r in _rows(wb["Curriculum"]):
        s.curriculum.append(dict(
            class_id=r["class_id"],
            subject_id=r["subject_id"],
            hours=_int(r.get("hours")),
            teacher_id=r.get("teacher_id", ""),
            blocks=r.get("blocks", ""),
            # How many groups this class splits into for THIS subject.
            # Default 1 - never assume a split. Proved necessary by last
            # year's file: 4رياضيات1 ran Computer Science whole-class while
            # splitting for Natural Sciences and Physics.
            groups=max(1, _int(r.get("groups"), 1)),
            room_type=r.get("room_type", ""),
            # S19, circular III.2: core / stream-specific subjects get three
            # quarters of their hours in the morning. yes = this row is one.
            core=(r.get("core") or "").strip().lower(),
            # Week A/B (the ministry fortnight patterns, T42): blank = every
            # week; A or B = that week of the two-week cycle only. hours are
            # the hours IN that week. Older workbooks have no such column.
            week=(r.get("week") or "").strip().upper(),
        ))
    # H14 Options sheet: one row per option GROUP (a teacher + the pupils
    # who chose that option, pooled from the listed classes). Bands are
    # derived: groups sharing any class must run simultaneously.
    if "Options" in wb.sheetnames:
        for r in _rows(wb["Options"]):
            s.options.append(dict(
                id=r["id"],
                subject_id=(r.get("subject_id") or "").strip(),
                teacher_id=(r.get("teacher_id") or "").strip(),
                hours=_int(r.get("hours")),
                blocks=(r.get("blocks") or "").strip(),
                classes=[c.strip() for c in (r.get("classes") or "")
                         .replace(",", ";").split(";") if c.strip()],
                room_type=(r.get("room_type") or "").strip(),
                # optional band column: rows sharing a band id run
                # SIMULTANEOUSLY; an EMPTY cell = this row is its own band
                # (a joint lesson - e.g. pooled sport). NO band column at
                # all = the H14 default: choice options sharing a class are
                # auto-linked into one simultaneous band.
                band=((r.get("band") or "").strip()
                      if "band" in r else None),
            ))
    compute_option_bands(s)

    # The Weights sheet is Majd's editable copy of the rule weights: any row
    # there OVERRIDES config.json. His workbook is the database - he asked
    # to see and change the weights without touching JSON (2026-08-25).
    if "Weights" in wb.sheetnames:
        n_over = 0
        for r in _rows(wb["Weights"]):
            key = (r.get("key") or "").strip()
            raw = str(r.get("value", "")).strip()
            if key and raw:
                # Majd 2026-08-25 ("what if we stuff more hard rules"):
                # HARD / صارم promotes a comfort rule to an unbreakable one.
                # The price: with too many HARD rules no timetable may exist
                # at all - phase 1 will say so plainly.
                if raw.upper() in ("HARD", "صارم"):
                    cfg.weights[key] = "HARD"
                else:
                    cfg.weights[key] = _int(raw, cfg.weights.get(key, 0))
                n_over += 1
        if n_over:
            print("NOTE : %d rule weights read from the Weights sheet "
                  "(they override config.json)." % n_over)
    if "Unavailable" in wb.sheetnames:
        for r in _rows(wb["Unavailable"]):
            s.unavailable.append(dict(
                teacher_id=r["teacher_id"],
                day=r.get("day", "*") or "*",
                period=r.get("period", "*") or "*",
                hard=(r.get("hard") or "yes").lower(),
                reason=r.get("reason", ""),
            ))
    if "Locked" in wb.sheetnames:
        for r in _rows(wb["Locked"]):
            s.locked.append(dict(
                class_id=r["class_id"],
                subject_id=r["subject_id"],
                day=r["day"],
                period=_int(r.get("period")),
                room_id=r.get("room_id", ""),
                why=r.get("why", ""),
            ))
    wb.close()
    return s


def compute_option_bands(s):
    """Group the option groups into BANDS: groups sharing any class must run
    simultaneously (Majd: while options run, a pupil can only be in another
    option - nobody misses a lesson). Idempotent; safe on empty options."""
    s.option_bands = []
    if not s.options:
        return s.option_bands
    parent = {}

    def find(a):
        while parent.get(a, a) != a:
            parent[a] = parent.get(parent[a], parent[a])
            a = parent[a]
        return a

    def union(a, b):
        parent.setdefault(a, a)
        parent.setdefault(b, b)
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    explicit = any(g.get("band") is not None for g in s.options)
    if explicit:
        # explicit band column: same id = simultaneous; blank = solo band
        comps = {}
        solo = 0
        for g in s.options:
            b = g.get("band") or ""
            if not b:
                solo += 1
                b = "__solo%d" % solo
            comps.setdefault(b, []).append(g)
    else:
        by_class = {}
        for g in s.options:
            parent.setdefault(g["id"], g["id"])
            for c in g["classes"]:
                if c in by_class:
                    union(g["id"], by_class[c])
                by_class[c] = g["id"]
        comps = {}
        for g in s.options:
            comps.setdefault(find(g["id"]), []).append(g)
    for n, (root, groups) in enumerate(sorted(comps.items()), start=1):
        classes = sorted({c for g in groups for c in g["classes"]})
        s.option_bands.append(dict(
            id="B%d" % n,
            classes=classes,
            groups=groups,
            hours=groups[0]["hours"],
            blocks=groups[0]["blocks"],
        ))
    # the printable views need a name for the band's pseudo-subject
    for b in s.option_bands:
        sid = "OPT:" + b["id"]
        if sid not in s.subjects:
            s.subjects[sid] = dict(id=sid, name="حصة الخيارات", short="خيار",
                                   difficulty="medium", room_type="__opt__",
                                   latest_period=0)
    return s.option_bands


def option_room_type(s, g):
    """The room an option group needs: its own column, else its subject's."""
    return (g.get("room_type")
            or s.subjects.get(g["subject_id"], {}).get("room_type")
            or "normal")


def check(s):
    """Plain-language validation. Returns (errors, notes)."""
    errs, notes = [], []
    days = set(s.cfg.days)

    max_run = longest_open_run(s.cfg)
    for c in s.curriculum:
        where = "Curriculum row " + c["class_id"] + " / " + c["subject_id"]
        if c.get("week", "") not in ("", "A", "B", "ALT", "ALT2"):
            errs.append(where + " - week '" + c["week"] + "' must be blank "
                        "(every week), A, B, ALT or ALT2 (groups take turns; "
                        "ALT2 is the swap side of the TP carousel).")
        if c.get("week", "") in ("ALT", "ALT2") and max(1, c.get("groups", 1)) < 2:
            errs.append(where + " - week ALT means THE GROUPS take turns "
                        "(odd groups week A, even week B), so it needs "
                        "groups of 2 or more.")
        if c["class_id"] not in s.classes:
            errs.append(where + " - class '" + c["class_id"] + "' is not in the Classes sheet.")
        if c["subject_id"] not in s.subjects:
            errs.append(where + " - subject '" + c["subject_id"] + "' is not in the Subjects sheet.")
        if c["teacher_id"] and c["teacher_id"] not in s.teachers:
            errs.append(where + " - teacher '" + c["teacher_id"] + "' is not in the Teachers sheet.")
        if c["hours"] <= 0:
            errs.append(where + " - hours must be greater than 0.")
        # H9: an EXPLICIT block pattern must be readable and must fit the
        # week. A blank pattern imposes nothing - the solver places single
        # hours freely (spreading is then a soft rule, not a promise).
        if str(c.get("blocks", "")).strip():
            bl, berr = parse_blocks(c.get("blocks", ""), c["hours"])
            if berr:
                errs.append(where + " - " + berr)
            elif bl:
                if sum(bl) != c["hours"]:
                    errs.append(where + " - blocks " + str(c.get("blocks")) + " add up to " +
                                str(sum(bl)) + " but hours says " + str(c["hours"]) + ".")
                if len(bl) > len(s.cfg.days):
                    errs.append(where + " - " + str(len(bl)) + " blocks but the week only has " +
                                str(len(s.cfg.days)) + " days (each block goes on its own day, H9).")
                if max(bl) > max_run:
                    errs.append(where + " - a block of " + str(max(bl)) + " consecutive hours can "
                                "never be placed: the longest open run in any day is " +
                                str(max_run) + " (the lunch break interrupts every day).")

    # HARD-promoted weights: everything may be promoted except the keys where
    # "zero" would mean nonsense (extra_day_present=HARD would forbid a
    # compact teacher from coming to school at all).
    for k, v in s.cfg.weights.items():
        if v == "HARD" and k in ("extra_day_present",):
            errs.append("Weights sheet: '" + k + "' cannot be HARD - zero of "
                        "it would mean the teacher never comes in. Use a "
                        "number.")

    # H14 options: refs exist, same-year pooling, one band per class, and
    # every group of a band identical in hours+blocks (they run as one).
    class_band = {}
    for g in s.options:
        where = "Options row " + str(g["id"])
        if g["subject_id"] not in s.subjects:
            errs.append(where + " - subject '" + g["subject_id"] + "' is not in the Subjects sheet.")
        if g["teacher_id"] and g["teacher_id"] not in s.teachers:
            errs.append(where + " - teacher '" + g["teacher_id"] + "' is not in the Teachers sheet.")
        if g["hours"] <= 0:
            errs.append(where + " - hours must be greater than 0.")
        if not g["classes"]:
            errs.append(where + " - no classes listed; write them like C01;C02.")
        for c in g["classes"]:
            if c not in s.classes:
                errs.append(where + " - class '" + c + "' is not in the Classes sheet.")
        grades = {str(s.classes[c].get("grade", "")) for c in g["classes"]
                  if c in s.classes}
        if len(grades) > 1:
            errs.append(where + " - pools classes of different years (" +
                        ", ".join(sorted(grades)) + "). Majd's rule: same "
                        "year only (streams may mix).")
        if str(g.get("blocks", "")).strip():
            bl, berr = parse_blocks(g["blocks"], g["hours"])
            if berr:
                errs.append(where + " - " + berr)
            elif bl and sum(bl) != g["hours"]:
                errs.append(where + " - blocks add up to " + str(sum(bl)) +
                            " but hours says " + str(g["hours"]) + ".")
        rt = option_room_type(s, g)
        if not s.rooms_of_type(rt) and s.rooms:
            errs.append(where + " - needs a '" + rt + "' room but none exists.")
    for band in s.option_bands:
        if len({g["hours"] for g in band["groups"]}) > 1:
            errs.append("Option band " + band["id"] + " (" +
                        ", ".join(g["id"] for g in band["groups"]) +
                        ") mixes different hours - groups that share classes "
                        "run SIMULTANEOUSLY, so they need the same hours.")
        if len({str(g.get("blocks", "")).strip() for g in band["groups"]}) > 1:
            errs.append("Option band " + band["id"] + " mixes different "
                        "blocks patterns - they run simultaneously, so the "
                        "pattern must match.")
        seen_t = set()
        for g in band["groups"]:
            t = g["teacher_id"]
            if t and t in seen_t:
                errs.append("Option band " + band["id"] + ": teacher " + t +
                            " has TWO groups that must run at the same time - "
                            "impossible. Give one group another teacher.")
            seen_t.add(t)
        explicit_bands = any(g.get("band") is not None for g in s.options)
        for c in band["classes"]:
            if c in class_band:
                msg = ("Class " + c + " appears in two option bands (" +
                       class_band[c] + " and " + band["id"] + ") - every "
                       "pupil takes exactly ONE option, so one band per "
                       "class.")
                if explicit_bands:
                    # hand-declared (or extracted-from-reality) bands may
                    # legitimately stagger - note it, don't block
                    notes.append(msg + " Allowed because bands are explicit.")
                else:
                    errs.append(msg)
            class_band[c] = band["id"]

    for t in s.teachers.values():
        off = t["day_off"]
        if off and off not in days and off != "(none)":
            errs.append("Teacher " + t["id"] + " has day_off '" + off +
                        "' which is not a school day (" + ", ".join(s.cfg.days) + ").")
        tw = t.get("travels_with", "")
        if tw and tw not in s.teachers:
            errs.append("Teacher " + t["id"] + " travels_with '" + tw +
                        "' which is not a teacher id in the Teachers sheet.")
        if tw and tw == t["id"]:
            errs.append("Teacher " + t["id"] + " travels_with themselves - "
                        "name the OTHER teacher of the pair.")
        tr = t.get("training_day", "")
        if tr and tr not in days and tr != "(none)":
            errs.append("Teacher " + t["id"] + " has training_day '" + tr +
                        "' which is not a school day (" + ", ".join(s.cfg.days) + ").")
        if tr and tr == off and tr in days:
            notes.append("Teacher " + t["id"] + ": training_day equals day_off (" +
                         tr + "). Allowed, but check it is intended.")
        # H18 - inspector's note on the approved distribution sheet: the day
        # off must not sit next to the training day (training Thursday means
        # neither Wednesday nor Friday may be the day off). Majd: applies to
        # ALL teachers. This is a data property - no placement can fix it.
        if off in days and tr in days and off != tr:
            day_list = list(s.cfg.days)
            gap = abs(day_list.index(off) - day_list.index(tr))
            # Majd 2026-08-24: Sat + Mon also count as adjacent - the Sunday
            # rest day between them would make three free days in a row.
            if gap == 1 or gap == len(day_list) - 1:
                errs.append("Teacher " + t["id"] + ": day_off " + off +
                            " is right next to training_day " + tr +
                            " - consecutive free days, Sunday included (H18, "
                            "inspector's rule). Pick a non-adjacent day off.")

    needed = {s.room_type_for(c) for c in s.curriculum}
    have = {r["type"] for r in s.rooms.values()}
    for rt in sorted(needed - have):
        errs.append("Some lessons need a '" + rt + "' room but no room of that type exists.")

    # teacher workload vs contract. A row with groups=N is taught N times
    # (each group separately), so the teacher works hours x N. A week-A row
    # only loads week A - the binding number is the BUSIER week.
    def taught_hours(c):
        return c["hours"] * max(1, c.get("groups", 1))

    def week_hours(c):
        """(week A card-hours, week B card-hours) of one curriculum row."""
        g = max(1, c.get("groups", 1))
        h = c["hours"]
        wk = c.get("week", "")
        if wk == "ALT":         # odd groups week A, even groups week B
            return (h * ((g + 1) // 2), h * (g // 2))
        if wk == "ALT2":        # the swap side: odd groups week B
            return (h * (g // 2), h * ((g + 1) // 2))
        return (h * g if wk in ("", "A") else 0,
                h * g if wk in ("", "B") else 0)

    loadw = {}      # tid -> [week A hours, week B hours]
    for c in s.curriculum:
        if c["teacher_id"]:
            e = loadw.setdefault(c["teacher_id"], [0, 0])
            a, b = week_hours(c)
            e[0] += a
            e[1] += b
    for g in s.options:            # H14: option groups run every week
        if g["teacher_id"]:
            e = loadw.setdefault(g["teacher_id"], [0, 0])
            e[0] += g["hours"]
            e[1] += g["hours"]
    # H10 compares the AVERAGE of the two weeks - that is how the school's
    # own sheets count (a fortnightly hour appears there as 0.5, e.g. the
    # official 18.5). The BUSIER week still drives the physical fit checks.
    load = {tid: max(e) for tid, e in loadw.items()}
    for tid in sorted(loadw):
        t = s.teachers.get(tid)
        e = loadw[tid]
        if t and t["hours"] and e[0] + e[1] > 2 * t["hours"]:
            errs.append("Teacher " + tid + " (" + t["name"] + ") is given " +
                        str((e[0] + e[1]) / 2.0) + " hours a week on average "
                        "(fortnightly hours count half, as on the official "
                        "sheets) but the contract says " + str(t["hours"]) + ".")
    for tid, t in s.teachers.items():
        if tid not in load:
            notes.append("Teacher " + tid + " (" + t["name"] + ") teaches nothing yet.")

    # a teacher's free days (day off + training day) and the H17 daily cap of 6
    # teaching hours bound what any placement could ever deliver
    for tid, hrs in sorted(load.items()):
        t = s.teachers.get(tid)
        if not t:
            continue
        blocked = {d for d in (t["day_off"], t.get("training_day", ""))
                   if d and d != "(none)"}
        # H17: at most 6 teaching hours on any one day (circular II.2)
        avail = sum(min(len(s.cfg.day_slots(d)), 6)
                    for d in s.cfg.days if d not in blocked)
        # a BLANK day_off is a soft preference since 2026-08-25 (Majd:
        # "maybe let him teach then") - it no longer removes capacity.
        if hrs > avail:
            errs.append("Teacher " + tid + " needs " + str(hrs) + " hours but at most " +
                        str(avail) + " are reachable: max 6 per day (H17)" +
                        (", with " + " and ".join(sorted(blocked)) + " free" if blocked else "") + ".")

    # class workload vs week length - from the PUPIL's seat: a grouped row
    # still costs each pupil `hours` periods (their own group's session), and
    # a week-A row costs nothing in week B. The binding week decides.
    week = len(s.cfg.slots)
    clw = {}
    for c in s.curriculum:
        e = clw.setdefault(c["class_id"], [0, 0])
        wk = c.get("week", "")
        if wk in ("ALT", "ALT2"):
            # some group is busy those hours in BOTH weeks (each its own)
            e[0] += c["hours"]
            e[1] += c["hours"]
        else:
            if wk in ("", "A"):
                e[0] += c["hours"]
            if wk in ("", "B"):
                e[1] += c["hours"]
    for band in s.option_bands:    # the band occupies each member class
        for c in band["classes"]:
            e = clw.setdefault(c, [0, 0])
            e[0] += band["hours"]
            e[1] += band["hours"]
    for cid in sorted(clw):
        if max(clw[cid]) > week:
            errs.append("Class " + cid + " needs " + str(max(clw[cid])) +
                        " hours in its busier week but the week only has " +
                        str(week) + " open periods.")

    # H16: the ministry says do not split a class of 24 pupils or fewer
    for c in s.curriculum:
        if c.get("groups", 1) > 1:
            size = s.classes.get(c["class_id"], {}).get("size", 0)
            if size and size <= 24:
                notes.append("Class " + c["class_id"] + " has " + str(size) +
                             " pupils but " + c["subject_id"] + " is set to split into " +
                             str(c["groups"]) + " groups. The ministry says not to split at "
                             "24 or fewer (H16). Set groups=1 unless there is a reason.")

    # H15: a daylight limit that no timetable could satisfy
    for sid, sub in s.subjects.items():
        lp = sub.get("latest_period") or 0
        if not lp:
            continue
        if lp > s.cfg.periods_per_day:
            errs.append("Subject " + sid + " has latest_period " + str(lp) +
                        " but the day only has " + str(s.cfg.periods_per_day) +
                        " periods.")
            continue
        allowed = sum(1 for (d, p) in s.cfg.slots if p <= lp)
        need = max(
            sum(week_hours(c)[i] for c in s.curriculum
                if c["subject_id"] == sid)
            for i in (0, 1))
        n_rooms = len(s.rooms_of_type(sub.get("room_type") or "normal"))
        cap = allowed * max(1, n_rooms)
        if need > cap:
            errs.append("Subject " + sid + " needs " + str(need) +
                        " hours but may only use periods 1-" + str(lp) +
                        ", giving " + str(cap) + " slots.")

    # THE bottleneck: rooms. Grouped rows occupy a room PER GROUP; week A/B
    # rows only occupy their week - the busier week is what must fit.
    total = (max(sum(week_hours(c)[i] for c in s.curriculum)
                 for i in (0, 1)) if s.curriculum else 0) \
        + sum(g["hours"] for g in s.options)
    cap = len(s.rooms) * week
    if cap:
        use = 100.0 * total / cap
        line = ("Room load: " + str(total) + " lesson-hours to place into " +
                str(len(s.rooms)) + " rooms x " + str(week) + " periods = " +
                str(cap) + " room-slots (" + ("%.0f" % use) + "% full).")
        if total > cap:
            errs.append(line + "  IMPOSSIBLE - more lessons than room-slots exist.")
        elif use > 90:
            notes.append(line + "  Above 90% is extremely tight - expect poor results.")
        else:
            notes.append(line)

    # per-room-type bottleneck (again per week, groups counted)
    by_type = {}
    for c in s.curriculum:
        rt = s.room_type_for(c)
        e = by_type.setdefault(rt, [0, 0])
        a, b = week_hours(c)
        e[0] += a
        e[1] += b
    for g in s.options:            # every option group needs its own room
        rt = option_room_type(s, g)
        e = by_type.setdefault(rt, [0, 0])
        e[0] += g["hours"]
        e[1] += g["hours"]
    for rt in sorted(by_type):
        n = len(s.rooms_of_type(rt))
        if n and max(by_type[rt]) > n * week:
            errs.append("Rooms of type '" + rt + "': need " + str(max(by_type[rt])) +
                        " hours but only " + str(n) + " room(s) x " + str(week) +
                        " periods = " + str(n * week) + " available.")

    # a class must split into the SAME halves everywhere: rows of one class
    # that split must agree on the group count, or "group 1" would mean two
    # different sets of pupils in two subjects.
    gcount = {}
    for c in s.curriculum:
        g = max(1, c.get("groups", 1))
        if g > 1:
            prev = gcount.setdefault(c["class_id"], g)
            if prev != g:
                errs.append("Class " + c["class_id"] + " splits into " +
                            str(prev) + " groups in one subject and " + str(g) +
                            " in another. All split rows of one class must "
                            "use the same group count.")
    return errs, notes


def main():
    cfg = load_config()
    path = sys.argv[1] if len(sys.argv) > 1 else None
    s = load_school(path, cfg)
    print("Loaded: %d teachers, %d classes, %d rooms, %d subjects, %d curriculum rows"
          % (len(s.teachers), len(s.classes), len(s.rooms), len(s.subjects), len(s.curriculum)))
    print("Week: %d open periods (%d days)" % (len(s.cfg.slots), len(s.cfg.days)))
    errs, notes = check(s)
    for n in notes:
        print("  NOTE  :", n)
    for e in errs:
        print("  ERROR :", e)
    print("")
    print("DATA OK" if not errs else str(len(errs)) + " PROBLEM(S) - fix these first.")
    return 1 if errs else 0


if __name__ == "__main__":
    sys.exit(main())
