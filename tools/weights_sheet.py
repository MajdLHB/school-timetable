# -*- coding: utf-8 -*-
"""Put the rule weights INTO data/school.xlsx as a 'Weights' sheet.

Majd asked (2026-08-25): "can i see the weight of rules and modify them,
put that on db for me so i can check it and edit". His workbook IS the
database, so: one row per weight, with the rule it belongs to and a plain
Arabic explanation. data.py reads this sheet on every run and it OVERRIDES
config.json. Deleting a row (or blanking its value) falls back to
config.json. Re-running this script refreshes descriptions but KEEPS the
values already in the sheet.

    python tools/weights_sheet.py
"""
import json
import os
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(HERE, "data", "school.xlsx")
CFG = os.path.join(HERE, "config.json")

# key -> (rule, Arabic explanation). Bigger number = the solver fights
# harder for that rule when rules compete. 0 = the rule is switched off.
INFO = {
    "teacher_gap":              ("S1", "ثقوب في يوم الأستاذ (ساعة فراغ محشورة بين حصتين)"),
    "one_hour_day":             ("S2", "أستاذ يتنقل للمعهد من أجل ساعة واحدة في اليوم"),
    "class_gap":                ("S7", "ثقوب في يوم التلاميذ"),
    "class_one_hour_session":   ("S15", "قسم يحضر لساعة واحدة يتيمة في الصباح أو المساء"),
    "hard_subject_evening":     ("S3", "مادة صعبة في المساء بدل الصباح"),
    "morning_evening_imbalance": ("S5", "عدم التوازن بين حصص الصباح والمساء للأستاذ"),
    "same_subject_twice_a_day": ("S6", "تكديس ساعات نفس المادة في يوم واحد (للصفوف بلا نمط حصص)"),
    "same_subject_adjacent_days": ("S6", "نفس المادة في يومين متتاليين"),
    "late_subject":             ("S16", "مادة بعد الساعة المحددة لها (رياضيات بعد 16h، فيزياء 17-18h)"),
    "last_period":              ("S14", "أي حصة في الساعة الأخيرة 17-18"),
    "not_after":                ("S18", "مادة مباشرة بعد مادة ممنوعة (فلسفة بعد رياضة)"),
    "same_nature_adjacent":     ("S4", "مادتان من نفس الطبيعة متتاليتان (أدبية/علمية/اجتماعية)"),
    "core_morning":             ("S19", "ساعات المواد الأساسية خارج الصباح (المطلوب: 3/4 صباحا)"),
    "last_period_fairness":     ("S10", "أستاذ محشور في الساعة الأخيرة أكثر من يومين في الأسبوع"),
    "overloaded_day":           ("S8", "أكثر من 4 ساعات تدريس لأستاذ في يوم واحد (توزيع الأسبوع)"),
    "extra_day_present":        ("S8", "يوم حضور إضافي لأستاذ compact=yes (يفضل أيام أقل)"),
    "daylight_not_morning":     ("S12", "مادة نهارية (رياضة) خارج الصباح"),
    "bac_friday_evening":       ("S13", "قسم باكالوريا يدرس مساء الجمعة"),
    "travel_pair":              ("S21", "زميلا نقل مشترك يحضران في أيام مختلفة"),
    "bac_no_free_afternoon":    ("S17", "قسم باكالوريا بلا أي مساء حر من الاثنين إلى الخميس"),
    "tp_groups_same_day":       ("S22", "فوجا نفس الحصة التطبيقية في يومين مختلفين (المطلوب: متتاليان في نفس اليوم)"),
    "pupil_day_over6":          ("T26", "تلميذ يدرس أكثر من 6 ساعات في اليوم"),
    "pupil_half_over4":         ("T27", "تلميذ يدرس أكثر من 4 ساعات في نصف اليوم"),
    "not_same_day":             ("T37", "مادتان ممنوع جمعهما في نفس اليوم (تاريخ/جغرافيا)"),
    "double_not_at_start":      ("T41", "حصة مزدوجة لا تبدأ في أول نصف اليوم"),
}

HINT = ("bigger = the solver fights harder for it; 0 = rule off. "
        "This sheet OVERRIDES config.json. Delete a row to fall back.")


def main():
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    weights = json.load(open(CFG, encoding="utf-8"))["weights"]
    wb = load_workbook(XLSX)
    existing = {}
    if "Weights" in wb.sheetnames:
        ws = wb["Weights"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[0]:
                existing[str(row[0]).strip()] = row[1]
        wb.remove(ws)
    ws = wb.create_sheet("Weights")
    ws.sheet_view.rightToLeft = False
    ws.append(["key", "value", "rule", "الشرح - ماذا يعاقب هذا الوزن"])
    ws.append([HINT, "", "", "عدّل العمود value فقط ثم احفظ الملف"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
    for c in ws[2]:
        c.font = Font(italic=True, color="808080")
    for key in sorted(weights, key=lambda k: (INFO.get(k, ("Z",))[0], k)):
        rule, desc = INFO.get(key, ("?", ""))
        val = existing.get(key, weights[key])
        ws.append([key, val, rule, desc])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 70
    wb.save(XLSX)
    kept = sum(1 for k in weights if k in existing
               and existing[k] != weights[k])
    print("Weights sheet written: %d weights, %d of them keeping your "
          "edited values." % (len(weights), kept))


if __name__ == "__main__":
    main()
