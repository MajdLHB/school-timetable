# -*- coding: utf-8 -*-
"""Generate the Curriculum sheet from the Distribution sheet + curriculum.json.

Sources of truth:
  - rules/curriculum.json  : WHAT each stream studies (hours, session patterns)
  - Distribution sheet     : WHO teaches each stream, how many sections each
  - Classes sheet          : the real 41 classes with grade + stream

Documented first-draft policies (every deviation is written to
data/CURRICULUM_REPORT.md so nothing is silent):

  P1  (UPGRADED 2026-08-24, both aSc probes passed) group sessions are real
      groups=2 rows at per-group length; fortnightly sessions are real
      week=A/B rows. Which week a class takes is set by class parity within
      its stream, so weeks A and B stay balanced school-wide.
  P2  blocks patterns come straight from the circular's session lengths,
      per row (theory, TP and fortnight rows each carry their own pattern).
  P3  The Distribution gives section COUNTS per stream, not class numbers.
      Classes are paired to teachers deterministically: classes in id order,
      Distribution rows in sheet order. This pairing is a FREE CHOICE -
      swap classes between two teachers of the same subject+stream freely.
  P4  SPORT rows use the exact class lists in the Distribution notes
      column (synthetic test data, FLAG-8). TASH uses the normal count
      columns from the official art sheet; only opted pupils attend.
  P5  Options (ESP/ALL/ITA, MUS) are POOLED cross-class groups (H14, not
      built) - skipped entirely, listed in the report.
  P6  CS-stream IT splits by the component column: خ = ALGO, تك = ICT plus
      NET (3rd) / DB (4th). Subject rows ALGO/ICT/NET/DB are added to the
      Subjects sheet if missing (room type 'it').
  P7  TECHSCI TECH rows get a BLANK teacher (FLAG-6: the MECH/ELEC teacher
      model is unresolved) - the lessons are placed, the teacher is decided
      later by Majd.
  P8  core=yes on the circular's stream-defining subjects (1st year:
      Arabic/French/Maths). Adjustable.
  P9  (Majd, 2026-08-24) lab subjects are SPLIT: the whole-class sessions
      become a theory row in an ORDINARY classroom, the group sessions
      become a <SID>_TP row in the lab, at per-group length (the two
      groups run back to back - ministry M-SN4). IT and TECH keep their
      dedicated rooms for everything ("except IT always in a lab").

Refuses to run if the Curriculum sheet already has data.

    python tools/make_curriculum.py
"""
import json
import math
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# an explicit path lets a regeneration be rehearsed on a COPY first
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "data", "school.xlsx")
CURJ = os.path.join(HERE, "rules", "curriculum.json")
REPORT = os.path.join(HERE, "data", "CURRICULUM_REPORT.md")

# Distribution column header -> (grade, stream code used in the Classes sheet)
COLMAP = {
    "1ث ج مشترك": (1, "COMMON"),
    "2آداب": (2, "LETTERS"), "2علوم": (2, "SCIENCES"),
    "2اقتصاد وخدمات": (2, "ECONOMY"), "2تكنلوجية اعلامية": (2, "IT_TECH"),
    "3آداب": (3, "LETTERS"), "3رياضيات": (3, "MATHS"),
    "3علوم تجريبية": (3, "EXPSCI"), "3علوم تقنية": (3, "TECHSCI"),
    "3اقتصاد وتصرف": (3, "ECONOMY"), "3علوم اعلامية": (3, "CS"),
    "4آداب": (4, "LETTERS"), "4رياضيات": (4, "MATHS"),
    "4علوم تجريبية": (4, "EXPSCI"), "4علوم تقنية": (4, "TECHSCI"),
    "4اقتصاد وتصرف": (4, "ECONOMY"), "4علوم اعلامية": (4, "CS"),
}
STREAM_KEY = {"LETTERS": "Y3_Y4_LETTERS", "MATHS": "Y3_Y4_MATH",
              "EXPSCI": "Y3_Y4_EXPSCI", "TECHSCI": "Y3_Y4_TECHSCI",
              "CS": "Y3_Y4_CS", "ECONOMY": "Y3_Y4_ECO"}
OPTION_SUBJECTS = {"ESP", "ALL", "ITA", "MUS", "LANG3"}
# P8: the circular's stream-defining subjects (III.2)
CORE = {
    (1, "COMMON"): {"ARAB", "FREN", "MATH"},
    (3, "MATHS"): {"MATH"}, (4, "MATHS"): {"MATH"},
    (3, "EXPSCI"): {"SVT", "PHYS"}, (4, "EXPSCI"): {"SVT", "PHYS"},
    (3, "TECHSCI"): {"TECH"}, (4, "TECHSCI"): {"TECH"},
    (3, "CS"): {"ALGO", "ICT"}, (4, "CS"): {"ALGO", "ICT"},
    (3, "ECONOMY"): {"ECO", "GEST"}, (4, "ECONOMY"): {"ECO", "GEST"},
    (3, "LETTERS"): {"ARAB", "PHIL"}, (4, "LETTERS"): {"ARAB", "PHIL"},
}


def node_for(streams, grade, stream):
    """The curriculum.json subject table for one (grade, stream)."""
    if grade == 1:
        return streams["Y1_COMMON"]["subjects"]
    if grade == 2:
        return streams["Y2"]["by_stream"][stream]["subjects"]
    return streams[STREAM_KEY[stream]]["Y%d" % grade]


def parse_sessions(node):
    """P9 + T42/T43 (2026-08-24, both aSc probes passed): read the circular's
    session list into real machinery.

    theory  = whole-class sessions -> ordinary classroom (lab subjects) or
              the subject's own room.
    tp      = group sessions, groups=2, at PER-GROUP length: a circled(4)
              is 2h per group - the teacher teaches both, back to back in
              the lab (ministry M-SN4, solver rule S22).
    *_alt   = fortnightly (Week A/B) versions of the same. WHICH week each
              class takes is decided by class parity, so the A and B loads
              balance across the school.
    pairs   = alt_whole1_group4: 1h whole class one week, 2h per group the
              other week - the two halves MUST take opposite weeks.
    """
    flags = []
    out = dict(theory=[], theory_alt=[], tp=[], tp_alt=[], pairs=[])

    def per_group(total):
        pg = total / 2.0
        rounded = int(math.ceil(pg))
        if rounded != pg:
            flags.append("group session of %sh -> %.1fh per group, rounded up "
                         "to %dh (aSc cannot show half hours - same convention "
                         "the school already used)" % (total, pg, rounded))
        return rounded

    for se in node.get("sessions", []):
        keys = list(se.keys())
        if keys == ["whole"]:
            out["theory"].append(int(se["whole"]))
        elif keys == ["fortnight_whole"]:
            out["theory_alt"].append(int(math.ceil(float(se["fortnight_whole"]))))
        elif keys == ["group"]:
            out["tp"].append(per_group(se["group"]))
        elif keys == ["fortnight_group"]:
            out["tp_alt"].append(per_group(se["fortnight_group"]))
        elif keys == ["alt_whole_1_2"]:
            # 1h one week, 2h the next = 1h weekly + a fortnightly extra hour
            out["theory"].append(1)
            out["theory_alt"].append(1)
        elif keys == ["alt_whole1_group4"]:
            # week X: 1h whole class; week Y: 2h per group - opposite weeks
            out["pairs"].append((1, 2))
        else:
            flags.append("session %r not understood - skipped" % (se,))
    return out, flags


def class_names_in(text, name_to_id):
    """Find real class names inside a free-text note."""
    found = []
    for name, cid in name_to_id.items():
        if name and name in (text or ""):
            found.append((name, cid))
    return [cid for _n, cid in sorted(found)]


def main():
    from openpyxl import load_workbook

    streams = json.load(open(CURJ, encoding="utf-8"))["streams"]
    wb = load_workbook(XLSX)
    cur_ws = wb["Curriculum"]
    if cur_ws.max_row > 2:
        sys.exit("REFUSING: the Curriculum sheet already has data. Clear it "
                 "first if you want a regeneration.")

    # ---- classes by (grade, stream) --------------------------------------
    classes = []            # (id, name, grade, stream)
    rows = list(wb["Classes"].iter_rows(values_only=True))
    header = [str(h or "").strip() for h in rows[0]]
    for r in rows[2:]:
        rec = dict(zip(header, r))
        if not rec.get("id"):
            continue
        classes.append((str(rec["id"]), str(rec.get("name") or ""),
                        int(float(rec.get("grade") or 0)),
                        str(rec.get("stream") or "").strip()))
    by_gs = {}
    name_to_id = {}
    for cid, name, grade, stream in classes:
        by_gs.setdefault((grade, stream), []).append(cid)
        name_to_id[name] = cid
    for k in by_gs:
        by_gs[k].sort()

    # ---- distribution rows ------------------------------------------------
    drows = list(wb["Distribution"].iter_rows(values_only=True))
    dheader = [str(h or "").strip() for h in drows[0]]
    dist = []
    for r in drows[2:]:
        rec = dict(zip(dheader, r))
        if rec.get("subject"):
            dist.append(rec)

    # ---- existing subjects + their room types -----------------------------
    subj_ws = wb["Subjects"]
    srows = list(subj_ws.iter_rows(values_only=True))
    sh = [str(h or "").strip() for h in srows[0]]
    subj_room = {}
    subj_name = {}
    for r in srows[2:]:
        rec = dict(zip(sh, r))
        if rec.get("id"):
            subj_room[str(rec["id"]).strip()] = str(rec.get("room_type") or "").strip()
            subj_name[str(rec["id"]).strip()] = str(rec.get("name") or "").strip()
    subj_ids = set(subj_room)
    CS_SUBJ = {"ALGO": "خوارزميات وبرمجة", "ICT": "تكنولوجيات المعلومات",
               "NET": "الشبكات والأنظمة", "DB": "قواعد البيانات"}
    added_subjects = []
    for sid, name_ar in CS_SUBJ.items():
        if sid not in subj_ids:
            subj_ws.append([sid, name_ar, sid, "", "it"])
            subj_ids.add(sid)
            subj_room[sid] = "it"
            added_subjects.append(sid)

    # P9: subjects whose sheet room type is a LAB get split into theory
    # (normal classroom) + a separate <SID>_TP subject in the lab. IT and
    # TECH families keep their dedicated rooms for everything, as Majd said
    # ("except IT always in a lab") and as the ministry says for technology.
    LAB_TYPES = {"lab_phys", "lab_sci", "lab_chem"}
    lab_split = {sid: rt for sid, rt in subj_room.items() if rt in LAB_TYPES}
    for sid, rt in sorted(lab_split.items()):
        tp_id = sid + "_TP"
        if tp_id not in subj_ids:
            subj_ws.append([tp_id, "أشغال تطبيقية - " + (subj_name.get(sid) or sid),
                            "ت.ط", "", rt])
            subj_ids.add(tp_id)
            subj_room[tp_id] = rt
            added_subjects.append(tp_id)

    report = ["# Curriculum generation report",
              "",
              "Generated by tools/make_curriculum.py. The policies P1-P8 are",
              "documented at the top of that file. EVERYTHING unusual is",
              "listed below - if a row is not listed, it came straight from",
              "the circular and the official distribution.",
              ""]
    R = report.append

    out_rows = []          # (class_id, subject_id, hours, teacher, blocks, core)
    flags_hours = set()
    unassigned = []        # (class, subject) with no teacher
    skipped = []           # option rows skipped

    # ---- P3: assign sections to classes, per subject ----------------------
    # assign[(class, subject)] = teacher_id
    assign = {}
    over = []
    # cursor per (grade, stream, subject) walking the sorted class list
    cursor = {}
    for rec in dist:
        subj = str(rec["subject"]).strip()
        tid = str(rec.get("teacher_id") or "").strip()
        comp = str(rec.get("component") or "").strip()
        notes = str(rec.get("notes") or "")
        if subj in OPTION_SUBJECTS:
            skipped.append((subj, tid, rec.get("hours_assigned")))
            continue
        if subj == "SPORT":
            # P4: the exact class lists live in the notes column (FLAG-8)
            for cid in class_names_in(notes, name_to_id):
                assign[cid, "SPORT"] = tid
            continue
        # TASH flows through the normal count columns like everything else
        # (the official art sheet fills them; the notes are just a summary)
        if subj in ("MECH", "ELEC"):
            continue       # P7 - reported below via TECH rows
        for col, (grade, stream) in COLMAP.items():
            n = rec.get(col)
            if not n:
                continue
            n = int(float(n))
            pool = by_gs.get((grade, stream), [])
            # CS-stream IT rows split into components
            targets = [subj]
            if subj == "IT" and stream == "CS":
                if comp == "خ":
                    targets = ["ALGO"]
                elif comp == "تك":
                    targets = ["ICT", "NET"] if grade == 3 else ["ICT", "DB"]
            for t_subj in targets:
                cur = cursor.get((grade, stream, t_subj), 0)
                take = pool[cur:cur + n]
                if len(take) < n:
                    over.append("%s %s: %s wants %d sections of %s but only %d "
                                "classes remain unassigned" %
                                (subj, tid, col, n, stream, len(take)))
                for cid in take:
                    assign[cid, t_subj] = tid
                cursor[(grade, stream, t_subj)] = cur + len(take)

    # ---- build the curriculum rows from curriculum.json -------------------
    guessed_it = []
    for cid, cname, grade, stream in sorted(classes):
        try:
            table = node_for(streams, grade, stream)
        except KeyError:
            R("- **%s (%s)**: no curriculum table for grade %d stream %s - "
              "NO ROWS generated. Fix the stream code." % (cid, cname, grade, stream))
            continue
        subs = dict(table)
        # f7: 2علوم/2آداب IT exists on the official sheets (~2h) but not in
        # curriculum.json. If a teacher was assigned, emit a GUESSED row.
        if (cid, "IT") in assign and "IT" not in subs:
            subs["IT"] = {"pupil_hours": 2.0, "sessions": [{"group": 4}]}
            guessed_it.append(cid)
        # Week A/B balance: even classes of a stream take their fortnightly
        # rows in week A, odd classes in week B - so both weeks carry about
        # half the school's fortnight load and no week is overloaded.
        pool = by_gs.get((grade, stream), [])
        parity = pool.index(cid) if cid in pool else 0
        prim = "A" if parity % 2 == 0 else "B"
        seco = "B" if prim == "A" else "A"
        for sid, node in subs.items():
            if sid.startswith("_"):
                continue
            if sid in OPTION_SUBJECTS:
                continue
            parts, hf = parse_sessions(node)
            for f in hf:
                flags_hours.add("%s: %s" % (sid, f))
            tid = assign.get((cid, sid), "")
            if not tid and sid == "TECH" and stream == "TECHSCI":
                pass       # P7, reported once below
            elif not tid and sid != "SPORT":
                unassigned.append((cid, sid))
            core = "yes" if sid in CORE.get((grade, stream), set()) else ""
            # P9: lab subjects put theory in an ordinary classroom and the
            # group sessions in the lab under <SID>_TP. Everything else
            # keeps its own subject id and room for both kinds of session.
            tp_sid = sid + "_TP" if sid in lab_split else sid
            theory_room = "normal" if sid in lab_split else ""
            rows_here = []   # (subject_id, hours, blocks, groups, room, week)
            if parts["theory"]:
                rows_here.append((sid, sum(parts["theory"]),
                                  "+".join(str(w) for w in
                                           sorted(parts["theory"], reverse=True)),
                                  1, theory_room, ""))
            for n in parts["theory_alt"]:
                rows_here.append((sid, n, str(n), 1, theory_room, prim))
            if parts["tp"]:
                rows_here.append((tp_sid, sum(parts["tp"]),
                                  "+".join(str(t) for t in
                                           sorted(parts["tp"], reverse=True)),
                                  2, "", ""))
            for n in parts["tp_alt"]:
                rows_here.append((tp_sid, n, str(n), 2, "", prim))
            for wh, tpg in parts["pairs"]:
                # 1h whole class one week, groups the other - opposite weeks
                rows_here.append((sid, wh, str(wh), 1, theory_room, prim))
                rows_here.append((tp_sid, tpg, str(tpg), 2, "", seco))
            for rsid, h, bl, g, rm, wk in rows_here:
                if h <= 0:
                    continue
                out_rows.append((cid, rsid, h, tid, bl, g, rm,
                                 core if rsid == sid else "", wk))
        # TASH: only for the classes on the official art sheet (opted pupils)
        if (cid, "TASH") in assign:
            out_rows.append((cid, "TASH", 2, assign[cid, "TASH"], "2",
                             1, "", "", ""))

    # ---- write ------------------------------------------------------------
    # make sure the sheet carries the week column (older workbooks stop at core)
    header = [str(c.value or "").strip() for c in cur_ws[1]]
    if "week" not in header:
        cur_ws.cell(row=1, column=len(header) + 1, value="week")
    for cid, sid, hours, tid, blocks, groups, room, core, week in out_rows:
        # Columns: class_id, subject_id, hours, teacher_id, blocks, groups,
        # room_type, core, week
        cur_ws.append([cid, sid, hours, tid, blocks, groups, room, core, week])
    wb.save(XLSX)

    # ---- report -----------------------------------------------------------
    R("## Numbers")
    R("")
    R("- %d curriculum rows written for %d classes" % (len(out_rows), len(classes)))
    R("- %d rows have a teacher; %d are BLANK (listed below)"
      % (sum(1 for r in out_rows if r[3]), sum(1 for r in out_rows if not r[3])))
    if added_subjects:
        R("- Subjects added for the CS stream (P6): " + ", ".join(added_subjects))
    R("")
    R("## P1 - session notes (rounding, unusual patterns)")
    R("")
    for f in sorted(flags_hours):
        R("- " + f)
    R("")
    R("## P3 - the class pairing is a FREE choice")
    R("")
    R("Classes were paired to teachers in id order, Distribution rows in")
    R("sheet order. The official sheets only fix the COUNTS per stream.")
    R("Swap classes between two teachers of the same subject freely.")
    if over:
        R("")
        R("**Count mismatches:**")
        for o in over:
            R("- " + o)
    R("")
    R("## Rows with a BLANK teacher")
    R("")
    R("These lessons ARE placed in the timetable, but nobody teaches them")
    R("yet - the solver applies no teacher rules to them. Decide and fill.")
    R("")
    tech_blank = sorted(set(c for c, s in unassigned if s == "TECH"))
    if tech_blank:
        R("- **TECH for %s** - FLAG-6: the MECH/ELEC teacher model is "
          "unresolved (each engineering teacher has both classes of one "
          "year at 16h). Majd decides." % ", ".join(tech_blank))
    for cid, sid in sorted(u for u in unassigned if u[1] != "TECH"):
        R("- %s / %s" % (cid, sid))
    R("")
    R("## P5 - options skipped (pooled groups, H14 not built)")
    R("")
    for subj, tid, h in skipped:
        R("- %s (teacher %s, %s h on the official sheet)" % (subj, tid, h))
    if guessed_it:
        R("")
        R("## GUESSED rows")
        R("")
        R("- IT 2h for %s - on the official sheets but not in the circular "
          "guide (curriculum.json). Confirm the hours." % ", ".join(guessed_it))
    R("")
    R("## Known carried-over guesses")
    R("")
    R("- ENGL T025: one 3ع-إعلا section moved to 3ع-تج in the "
      "Distribution (documented guess) - the pairing here inherits it.")
    R("- SPORT class lists are synthetic test data (FLAG-8).")

    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(report))
    print("Wrote %d curriculum rows." % len(out_rows))
    print("Report: %s" % REPORT)


if __name__ == "__main__":
    main()
