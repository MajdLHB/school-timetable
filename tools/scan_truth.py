# -*- coding: utf-8 -*-
"""Scan LAST YEAR's real file against the MINISTRY guide and write the
'how it is actually done' reference - Majd 2026-08-25: 'scan last year
data with some info from ministry and extract how its done correctly'.

Output: data/TP_TRUTH.md (private) - per stream and subject:
  - what last year really ran (whole hours, group hours, weekly/fortnight/
    carousel, session blocks)
  - what the ministry guide prescribes (curriculum.json sessions)
  - the observed CAROUSEL partners (whose groups really shared slots)
No teacher or pupil names appear - streams, subjects, hours only.

    python tools/scan_truth.py
"""
import collections
import glob
import json
import os
import sys
import xml.etree.ElementTree as ET

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "data", "TP_TRUTH.md")

# subject-name keyword -> guide id
NAME2ID = [
    ("فيزيائية", "PHYS"), ("الحياة", "SVT"), ("حياة", "SVT"),
    ("تقنية", "TECH"), ("علامية", "IT"), ("رياضيات", "MATH"),
    ("عربية", "ARAB"), ("فرنسية", "FREN"), ("نقليزية", "ENGL"),
    ("تاريخ", "HIST"), ("جغرافيا", "GEO"), ("تفكير", "PISL"),
    ("إسلامية", "PISL"), ("مدنية", "CIV"), ("فلسفة", "PHIL"),
    ("خوارزميات", "ALGO"), ("المعلومات", "ICT"), ("اقتصاد", "ECO"),
    ("تصرف", "GEST"), ("بدنية", "SPORT"), ("رياضة", "SPORT"),
    ("آلية", "MECH"), ("كهربائية", "ELEC"), ("إسبانية", "ESP"),
    ("ألمانية", "ALL"), ("يطالية", "ITA"), ("موسيق", "MUS"),
    ("تشكيلية", "TASH"),
]


def sub_id(name):
    for k, v in NAME2ID:
        if k in name:
            return v
    return None


def stream_of(cname):
    n = cname.strip()
    g = next((ch for ch in n if ch.isdigit()), "?")
    if "ث" in n and g == "1":
        return (1, "COMMON", "1ère commune")
    for kw, code, label in (
            ("آداب", "LETTERS", "آداب"), ("تجريبية", "EXPSCI", "علوم تجريبية"),
            ("تقنية", "TECHSCI", "علوم تقنية"), ("رياض", "MATHS", "رياضيات"),
            ("علامية", "CS", "علوم إعلامية"), ("اقتصاد", "ECONOMY", "اقتصاد"),
            ("تكنلوج", "IT_TECH", "تكنولوجية إعلامية"),
            ("علوم", "SCIENCES", "علوم")):
        if kw in n:
            return (int(g) if g.isdigit() else 0, code, "%s%s" % (g, label))
    return (int(g) if g.isdigit() else 0, "?", n[:8])


def guide_sessions(streams, grade, code, gid):
    try:
        if grade == 1:
            node = streams["Y1_COMMON"]["subjects"]
        elif grade == 2:
            node = streams["Y2"]["by_stream"][code]["subjects"]
        else:
            key = {"LETTERS": "Y3_Y4_LETTERS", "MATHS": "Y3_Y4_MATH",
                   "EXPSCI": "Y3_Y4_EXPSCI", "TECHSCI": "Y3_Y4_TECHSCI",
                   "CS": "Y3_Y4_CS", "ECONOMY": "Y3_Y4_ECO"}[code]
            node = streams[key]["Y%d" % grade]
        n = node.get(gid)
        if not n:
            return "—"
        return json.dumps(n.get("sessions", []), ensure_ascii=False)
    except KeyError:
        return "—"


def main():
    from openpyxl import load_workbook

    streams = json.load(open(os.path.join(HERE, "rules", "curriculum.json"),
                             encoding="utf-8"))["streams"]
    wb = load_workbook(os.path.join(HERE, "data", "school_lastyear.xlsx"),
                       read_only=True)
    subs = {str(r[0]): str(r[1] or "") for r in
            list(wb["Subjects"].iter_rows(values_only=True))[2:] if r and r[0]}
    cls = {str(r[0]): str(r[1] or "") for r in
           list(wb["Classes"].iter_rows(values_only=True))[2:] if r and r[0]}

    # aggregate rows per (stream, subject-name)
    agg = collections.defaultdict(lambda: collections.Counter())
    for r in list(wb["Curriculum"].iter_rows(values_only=True))[2:]:
        if not r or not r[0]:
            continue
        cid, sid, hours, tid, blocks, groups, rt, core, wk = (list(r) + [""])[:9]
        st = stream_of(cls.get(str(cid), ""))
        name = subs.get(str(sid), str(sid))
        kind = ("carousel(ALT)" if wk in ("ALT", "ALT2")
                else ("fortnight %s" % wk) if wk in ("A", "B")
                else "weekly")
        tag = ("%dh/group x%d groups, %s, blocks %s"
               % (hours, groups, kind, blocks or "-")) if (groups or 1) > 1 \
            else "%dh whole, %s, blocks %s" % (hours, kind, blocks or "-")
        agg[st, name][tag] += 1
    wb.close()

    # carousel partners: whose group sessions really shared slots
    hits = glob.glob(os.path.join(HERE, "data", "reference", "**", "*.xml"),
                     recursive=True)
    raw = open(hits[0], "rb").read().decode("cp1256", errors="replace")
    raw = raw.replace("windows-1252", "utf-8").replace("Windows-1252", "utf-8")
    root = ET.fromstring(raw)
    xsubs = {el.get("id"): el.get("name") or "" for el in root.iter("subject")}
    xcls = {el.get("id"): el.get("name") or "" for el in root.iter("class")}
    division = {el.get("id") for el in root.iter("group")
                if (el.get("entireclass") or "0") not in ("1", "true")}
    lessons = {el.get("id"): el for el in root.iter("lesson")}
    slotmap = collections.defaultdict(list)   # (class, day, period, wmask) -> subj names
    for el in root.iter("card"):
        L = lessons.get(el.get("lessonid"))
        if L is None:
            continue
        gids = (L.get("groupids") or "").split(",")
        if not any(g in division for g in gids):
            continue
        cids = [c for c in (L.get("classids") or "").split(",") if c]
        if len(cids) != 1:
            continue
        mask = el.get("days") or ""
        if mask.count("1") != 1:
            continue
        slotmap[cids[0], mask.index("1"), el.get("period"),
                el.get("weeks") or ""].append(xsubs.get(L.get("subjectid"), "?"))
    partners = collections.Counter()
    for key, names in slotmap.items():
        uniq = sorted(set(names))
        if len(uniq) > 1:
            st = stream_of(xcls.get(key[0], ""))
            for i in range(len(uniq)):
                for j in range(i + 1, len(uniq)):
                    partners[st[2], uniq[i], uniq[j]] += 1

    L = ["# كيف يتم فعليا - آخر سنة مقابل دليل الوزارة",
         "", "المصدر: ملف السنة الماضية الحقيقي + rules/curriculum.json.",
         "TP = الحصص التي جرت فعلا بالأفواج. لا أسماء هنا - مواد وأعداد فقط.", ""]
    A = L.append
    for (st, name) in sorted(agg, key=lambda k: (k[0][0], k[0][1], k[1])):
        pass
    cur_st = None
    for (st, name) in sorted(agg, key=lambda k: (k[0][0], k[0][2], k[1])):
        if st != cur_st:
            cur_st = st
            A("")
            A("## %s" % st[2])
            A("")
            A("| المادة | آخر سنة (الحقيقة) | دليل الوزارة (sessions) |")
            A("|---|---|---|")
        gid = sub_id(name)
        forms = "؛ ".join("%s ×%d قسم" % (t, n)
                          for t, n in sorted(agg[st, name].items()))
        A("| %s | %s | %s |"
          % (name, forms, guide_sessions(streams, st[0], st[1], gid) if gid else "—"))
    A("")
    A("## شراكات الأفواج المرصودة (فوج في مادة والفوج الآخر في مادة أخرى، نفس الحصة)")
    A("")
    for (stl, a, b), n in partners.most_common():
        A("- %s: **%s ↔ %s** (%d مرة)" % (stl, a, b, n))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print("wrote", OUT, "-", len(agg), "stream-subject entries,",
          len(partners), "carousel partnerships")


if __name__ == "__main__":
    main()
