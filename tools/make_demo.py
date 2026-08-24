"""Generate a FAKE school of roughly the real size, to test the tool.

No real person appears here. Teachers are 'Prof 01'. This exists so we can
measure how long solving actually takes at 40 classes / 20 rooms (the real school has 45 - see docs/RULES.md) before any
real data is loaded.

    python tools/make_demo.py            -> data/school.xlsx (40 classes)
    python tools/make_demo.py 20         -> a smaller 20-class school
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "solver"))

from openpyxl import load_workbook  # noqa: E402
import make_workbook  # noqa: E402

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "data", "school.xlsx")

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

# subject id, name, short, difficulty, room type, hours per class
SUBJECTS = [
    ("ARAB",  "Arabe",              "Ar", "medium", "normal",   3),
    ("FREN",  "Francais",           "Fr", "medium", "normal",   2),
    ("MATH",  "Mathematiques",      "Ma", "hard",   "normal",   3),
    ("PHYS",  "Sciences Physiques", "Ph", "hard",   "normal",   2),
    ("HIST",  "Histoire-Geo",       "HG", "medium", "normal",   2),
    ("SCI",   "Sciences Naturelles", "SN", "hard",  "lab_sci",  1),
    ("SPORT", "Education Physique", "EP", "easy",   "gym",      1),
]
# taught only to the lower half of the school, to make the data less uniform
IT = ("IT", "Informatique", "In", "medium", "it", 1)

ROOMS = [("normal", 15), ("lab_sci", 2), ("it", 1), ("gym", 2)]

MAX_TEACHER_HOURS = 18


def build(n_classes=40):
    # NEVER destroy real data. A demo workbook carries a hidden _DEMO sheet;
    # anything without it is real and must not be overwritten by fake names.
    if os.path.exists(OUT):
        from openpyxl import load_workbook as _lw
        try:
            _wb = _lw(OUT, read_only=True)
            _demo = "_DEMO" in _wb.sheetnames
            _wb.close()
        except Exception:
            _demo = False
        if not _demo:
            sys.exit("REFUSING: " + OUT + " holds REAL data (no _DEMO sheet). "
                     "Move it somewhere safe first if you really want a demo.")
        os.remove(OUT)
    make_workbook.build()
    wb = load_workbook(OUT)

    # ---------------- rooms ----------------
    rooms = []
    n = 0
    for rtype, count in ROOMS:
        for _ in range(count):
            n += 1
            rid = "R%02d" % n
            nice = {"normal": "Salle", "lab_sci": "Labo SVT",
                    "it": "Salle Info", "gym": "Gymnase"}[rtype]
            rooms.append((rid, "%s %d" % (nice, n), rtype, 35))
    ws = wb["Rooms"]
    for r in rooms:
        ws.append(list(r))

    # ---------------- subjects ----------------
    ws = wb["Subjects"]
    for sid, name, short, diff, rt, _h in SUBJECTS + [IT]:
        # Sport must finish by 16:00 - no stadium lighting (rule H15) - and
        # is exempt from the min-2h rules (circular I.2 note).
        latest = 8 if sid == "SPORT" else ""
        # S16 ministry preferences: Maths before 16:00 (M-MA3), Physics
        # avoids 17:00-18:00 (M-PH5).
        avoid = {"MATH": 8, "PHYS": 9}.get(sid, "")
        exempt = "yes" if sid == "SPORT" else ""
        ws.append([sid, name, short, diff, rt, latest, avoid, exempt])

    # ---------------- classes ----------------
    # Columns: id, name, grade, stream, size, is_bac, home_room, cohort -
    # matching the workbook header exactly.
    classes = []
    normal_rooms = [r[0] for r in rooms if r[2] == "normal"]
    for i in range(1, n_classes + 1):
        cid = "C%02d" % i
        grade = (i - 1) % 4 + 1
        stream = "COMMON" if grade == 1 else "SCIENCES"
        is_bac = "yes" if grade == 4 else ""
        classes.append((cid, "%d eme %d" % (grade, i), grade, stream, 30,
                        is_bac, normal_rooms[(i - 1) % len(normal_rooms)], "ALL"))
    ws = wb["Classes"]
    for c in classes:
        ws.append(list(c))

    # ---------------- curriculum + teachers ----------------
    # Work out how many hours of each subject the whole school needs, then
    # hire just enough teachers for each subject.
    plan = []           # (class_id, subject_id, hours)
    for cid, _n, grade, _st, _sz, _b, _r, _co in classes:
        for sid, _nm, _sh, _d, _rt, hours in SUBJECTS:
            plan.append((cid, sid, hours))
        if grade <= 2:
            plan.append((cid, IT[0], IT[5]))

    need = {}
    for cid, sid, h in plan:
        need[sid] = need.get(sid, 0) + h

    teachers = []
    pool = {}           # subject -> [teacher ids]
    tno = 0
    for sid in [s[0] for s in SUBJECTS] + [IT[0]]:
        count = max(1, -(-need[sid] // MAX_TEACHER_HOURS))   # ceil
        pool[sid] = []
        for _ in range(count):
            tno += 1
            tid = "T%03d" % tno
            day_off = DAYS[tno % len(DAYS)]
            # Columns: id, name, short, subjects, hours, day_off,
            # training_day, compact, notes - matching the workbook header.
            # A few compact=yes teachers so the S8 exception is exercised.
            compact = "yes" if tno % 10 == 0 else ""
            teachers.append((tid, "Prof %03d" % tno, "P%02d" % tno, sid,
                             MAX_TEACHER_HOURS, day_off, "", compact,
                             "demo data - not a real person"))
            pool[sid].append(tid)

    # hand each class+subject to the least loaded qualified teacher
    load = {t[0]: 0 for t in teachers}
    curriculum = []
    for cid, sid, h in plan:
        best = min(pool[sid], key=lambda t: (load[t], t))
        load[best] += h
        blocks = "+".join(["1"] * h)
        # Columns: class_id, subject_id, hours, teacher_id, blocks, groups,
        # room_type - matching the workbook header.
        curriculum.append((cid, sid, h, best, blocks, 1, ""))

    ws = wb["Teachers"]
    for t in teachers:
        ws.append(list(t))
    ws = wb["Curriculum"]
    for c in curriculum:
        ws.append(list(c))

    # a few unavailabilities, so the feature is exercised
    ws = wb["Unavailable"]
    for t in teachers[:5]:
        ws.append([t[0], "Mon", 1, "yes", "demo - arrives late"])

    # Stamp the workbook as fake, so check_privacy.py knows there is nothing
    # personal in it. Real workbooks never have this sheet.
    mark = wb.create_sheet("_DEMO")
    mark["A1"] = "This workbook contains generated fake data. No real person appears in it."
    mark.sheet_state = "hidden"

    wb.save(OUT)
    total = sum(c[2] for c in curriculum)
    print("Demo school written to %s" % OUT)
    print("  %d classes, %d rooms, %d teachers, %d subjects"
          % (len(classes), len(rooms), len(teachers), len(SUBJECTS) + 1))
    print("  %d lesson-hours per week to place" % total)
    print("  NOTE: fake names only. Never commit data/ - see docs/PRIVACY.md")


if __name__ == "__main__":
    build(int(sys.argv[1]) if len(sys.argv) > 1 else 40)
