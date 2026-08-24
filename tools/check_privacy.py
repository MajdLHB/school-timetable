"""Refuse to let real people's data escape into a published repository.

Run this BEFORE pushing anywhere, ever. It does three things:

  1. Fails if git is tracking anything under data/ or out/, or any file type
     that carries real data (pdf, xlsx, roz, csv, images).
  2. Reads the REAL names out of data/school.xlsx and greps every tracked
     file for them - catching a name accidentally pasted into a doc or a
     code comment.
  3. Warns if the git HISTORY ever contained such a file, because deleting a
     file does not remove it from history.

Exit code 0 = CLEAN. Anything else = do not publish.
"""
import os
import subprocess
import sys

# Windows consoles default to cp1252, which cannot encode Arabic and raises
# UnicodeEncodeError mid-print. Force UTF-8 so real names are printable.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BLOCKED_DIRS = ("data/", "out/")
BLOCKED_EXT = (".pdf", ".xlsx", ".xls", ".csv", ".roz", ".docx",
               ".jpg", ".jpeg", ".png")
# Official ministry circulars are published public documents and contain no
# personal data - they are the specification this tool implements. Everything
# under rules/ must still be READ before it is committed.
ALLOWED_PREFIXES = ("rules/",)
# names short enough to cause false matches are skipped
MIN_NAME_LEN = 4


def git(*args):
    try:
        r = subprocess.run(["git"] + list(args), cwd=HERE,
                           capture_output=True, text=True, timeout=60)
        return r.stdout.splitlines() if r.returncode == 0 else []
    except (OSError, subprocess.SubprocessError):
        return []


def real_names():
    """Every human-identifying string in the live data file.

    Returns (names, is_demo). A workbook made by tools/make_demo.py carries a
    hidden _DEMO sheet; there is no real person in it, so the name scan is
    skipped. Anything without that sheet is treated as real data.
    """
    path = os.path.join(HERE, "data", "school.xlsx")
    if not os.path.exists(path):
        return [], False
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], False
    names = set()
    wb = load_workbook(path, read_only=True, data_only=True)
    if "_DEMO" in wb.sheetnames:
        wb.close()
        return [], True
    for sheet, cols in (("Teachers", ("name", "notes")),
                        ("Unavailable", ("reason",))):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h else "" for h in rows[0]]
        for r in rows[2:]:
            for h, v in zip(header, r):
                if h in cols and v and len(str(v).strip()) >= MIN_NAME_LEN:
                    names.add(str(v).strip())
    wb.close()
    return sorted(names), False


def names_from_reference():
    """Harvest real names from any aSc XML sitting in data/reference/.

    Added after 2026-08-24, when a real export (101 teachers) was dropped in a
    folder outside data/ and was committed. File-level rules now block it, but
    we also want to catch any of those names pasted into a doc or a comment.
    """
    import re
    out = set()
    root = os.path.join(HERE, "data", "reference")
    if not os.path.isdir(root):
        return out
    for dirpath, _dirs, files in os.walk(root):
        for fn in files:
            if not fn.lower().endswith(".xml"):
                continue
            try:
                raw = open(os.path.join(dirpath, fn), "rb").read()
            except OSError:
                continue
            # aSc labels its export encoding="windows-1252" but writes Arabic
            # in windows-1256. Decoding as utf-8 silently DELETES every Arabic
            # name (verified 2026-08-24: 101 names -> 0). cp1256 recovers them.
            text = raw.decode("cp1256", errors="ignore")
            # Only PEOPLE. Matching every name= attribute also caught the
            # file's own header ("aSc Timetables 2012 XML") - a false alarm.
            for tag in ("teacher", "student"):
                pat = '<' + tag + '[^>]*name="([^"]{4,60})"'
                for m in re.findall(pat, text):
                    out.add(m.strip())
    return out


def main():
    problems = []
    tracked = git("ls-files")
    if not tracked:
        print("Not a git repository (or nothing tracked yet) - nothing to leak.")

    for f in tracked:
        low = f.lower()
        if low.startswith(BLOCKED_DIRS):
            problems.append("TRACKED REAL DATA: " + f)
        elif low.startswith(ALLOWED_PREFIXES):
            continue
        elif low.endswith(BLOCKED_EXT):
            problems.append("TRACKED DATA FILE TYPE: " + f)

    names, is_demo = real_names()
    names = sorted(set(names) | names_from_reference())
    if names:
        for f in tracked:
            p = os.path.join(HERE, f)
            if not os.path.isfile(p):
                continue
            try:
                with open(p, encoding="utf-8", errors="ignore") as fh:
                    text = fh.read()
            except OSError:
                continue
            for n in names:
                if n in text:
                    problems.append("REAL NAME %r found inside tracked file %s" % (n, f))

    # history check
    hist = set()
    for line in git("log", "--pretty=format:", "--name-only", "--all"):
        line = line.strip()
        if not line:
            continue
        low = line.lower()
        if low.startswith(ALLOWED_PREFIXES):
            continue
        if low.startswith(BLOCKED_DIRS) or low.endswith(BLOCKED_EXT):
            hist.add(line)
    if hist:
        problems.append(
            "GIT HISTORY once contained real data (%d file(s), e.g. %s). "
            "Deleting them now does NOT remove them from history. "
            "Start a fresh repository instead."
            % (len(hist), sorted(hist)[0]))

    print("")
    if is_demo:
        print("data/school.xlsx is DEMO data (hidden _DEMO sheet) - no real")
        print("names to scan for. File-level rules still enforced below.")
    print("Checked %d tracked files against %d real names." % (len(tracked), len(names)))
    print("")
    if problems:
        print("!!! NOT SAFE TO PUBLISH - %d problem(s) !!!" % len(problems))
        print("")
        for p in problems:
            print("  " + p)
        print("")
        print("See docs/PRIVACY.md.")
        return 1
    print("CLEAN - no personal data is tracked by git.")
    print("Safe to publish the code. data/ and out/ stay on this PC.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
