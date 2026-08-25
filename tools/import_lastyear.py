# -*- coding: utf-8 -*-
"""THE FINAL DUEL, Majd's design (2026-08-25): rebuild LAST YEAR's school
from its own aSc export, let the solver play the exact same match, and
compare with tools/score_table.py - same pupils, same teachers, same hours.

Reads data/reference/*.xml, writes data/school_lastyear.xlsx (PRIVATE - the
data/ firewall applies). The current school.xlsx is NOT touched.

Extraction = last year's file is the source of truth:
  - hours per (class, subject, group, week) = the cards actually given
  - groups from the division groups aSc recorded (224 subdivisions)
  - week A/B from the card week masks (145 fortnight cards)
  - blocks = the session run-lengths the school actually used
  - multi-class lessons (23) -> the Options sheet
  - teacher contract = ceil(average weekly card-hours), day off left
    flexible (their real fixed days are unknown - noted caveat)
Rooms are typed 'normal' (type fidelity is unknown; the duel measures
comfort, not room labels - caveat printed).

    python tools/import_lastyear.py
"""
import collections
import glob
import math
import os
import sys
import xml.etree.ElementTree as ET

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "data", "school_lastyear.xlsx")


def weeks_of(mask):
    m = (mask or "").strip()
    if m == "10":
        return ("A",)
    if m == "01":
        return ("B",)
    return ("A", "B")


def main():
    from openpyxl import Workbook

    hits = glob.glob(os.path.join(HERE, "data", "reference", "**", "*.xml"),
                     recursive=True)
    if not hits:
        sys.exit("No reference export in data/reference/")
    # T3: aSc 2013 exports are cp1256 mislabelled windows-1252 - decode by
    # hand or every Arabic name turns to mojibake.
    raw = open(hits[0], "rb").read().decode("cp1256", errors="replace")
    raw = raw.replace("windows-1252", "utf-8").replace("Windows-1252", "utf-8")
    root = ET.fromstring(raw)

    teachers = {}
    for el in root.iter("teacher"):
        teachers[el.get("id")] = dict(name=el.get("name") or el.get("short") or "",
                                      short=el.get("short") or "")
    classes = {}
    for el in root.iter("class"):
        classes[el.get("id")] = dict(name=el.get("name") or el.get("short") or "",
                                     short=el.get("short") or "")
    rooms = {}
    for el in root.iter("classroom"):
        rooms[el.get("id")] = dict(name=el.get("name") or el.get("short") or "")
    subjects = {}
    for el in root.iter("subject"):
        subjects[el.get("id")] = dict(name=el.get("name") or el.get("short") or "",
                                      short=el.get("short") or "")
    # groups: entireclass flag + per-class numbering of division groups
    entire = set()
    gnum = {}
    per_class_count = collections.Counter()
    for el in root.iter("group"):
        gid = el.get("id")
        if (el.get("entireclass") or "0") in ("1", "true", "True"):
            entire.add(gid)
        else:
            cid = el.get("classid")
            per_class_count[cid] += 1
            gnum[gid] = per_class_count[cid]

    lessons = {}
    for el in root.iter("lesson"):
        lessons[el.get("id")] = dict(
            subject=el.get("subjectid"),
            classes=[c for c in (el.get("classids") or "").split(",") if c],
            groups=[g for g in (el.get("groupids") or "").split(",") if g],
            teachers=[t for t in (el.get("teacherids") or "").split(",") if t],
        )

    # cards -> per (class, subject, teacher, group#, week): list of (day, p)
    placed = collections.defaultdict(list)
    opt_cards = collections.defaultdict(list)   # multi-class lessons
    skipped = 0
    for el in root.iter("card"):
        L = lessons.get(el.get("lessonid"))
        mask = el.get("days") or ""
        if not L or mask.count("1") != 1:
            skipped += 1
            continue
        d = mask.index("1")
        p = int(float(el.get("period")))
        wks = weeks_of(el.get("weeks"))
        tid = L["teachers"][0] if L["teachers"] else ""
        if len(L["classes"]) > 1:
            opt_cards[el.get("lessonid")].append((d, p, wks))
            continue
        cid = L["classes"][0] if L["classes"] else ""
        if not cid:
            skipped += 1
            continue
        g = 0
        for gid in L["groups"]:
            if gid in gnum:
                g = gnum[gid]
                break
        wk = wks[0] if len(wks) == 1 else ""
        placed[cid, L["subject"], tid, g, wk].append((d, p))

    # ---- curriculum rows --------------------------------------------------
    # merge the group copies: hours per group must match; groups = how many
    rows = []
    flags = []
    by_rowkey = collections.defaultdict(dict)   # (cid,subj,tid,wk) -> {g: slots}
    for (cid, subj, tid, g, wk), slots in placed.items():
        by_rowkey[cid, subj, tid, wk][g] = slots

    def blocks_of(slots):
        byday = collections.defaultdict(list)
        for d, p in slots:
            byday[d].append(p)
        runs = []
        for d, ps in byday.items():
            ps.sort()
            run = 1
            for a, b in zip(ps, ps[1:]):
                if b == a + 1:
                    run += 1
                else:
                    runs.append(run)
                    run = 1
            runs.append(run)
        return "+".join(str(r) for r in sorted(runs, reverse=True))

    # regroup per (class, subject, teacher) so the CAROUSEL can be seen:
    # group 1 in week A + group 2 in week B (last year's real SVT/TECH
    # linkage, visible in the exported grid) becomes ONE row week=ALT.
    per_cst = collections.defaultdict(dict)
    for (cid, subj, tid, wk), by_g in by_rowkey.items():
        per_cst[cid, subj, tid][wk] = by_g
    for (cid, subj, tid), by_wk in sorted(per_cst.items()):
        # whole-class parts, per week, as they were
        for wk, by_g in sorted(by_wk.items()):
            if 0 in by_g:
                slots = by_g[0]
                rows.append((cid, subj, len(slots), tid, blocks_of(slots),
                             1, "", "", wk))
        gpart = {wk: {g: sl for g, sl in by_g.items() if g}
                 for wk, by_g in by_wk.items()}
        gpart = {wk: d for wk, d in gpart.items() if d}
        a, b = gpart.get("A", {}), gpart.get("B", {})
        if set(a) == {1} and set(b) == {2} and len(a[1]) == len(b[2]):
            rows.append((cid, subj, len(a[1]), tid, blocks_of(a[1]),
                         2, "", "", "ALT"))
            gpart.pop("A"), gpart.pop("B")
        elif set(a) == {2} and set(b) == {1} and len(a[2]) == len(b[1]):
            rows.append((cid, subj, len(b[1]), tid, blocks_of(b[1]),
                         2, "", "", "ALT2"))
            gpart.pop("A"), gpart.pop("B")
        for wk, by_g in sorted(gpart.items()):
            counts = {g: len(sl) for g, sl in by_g.items()}
            h = max(counts.values())
            if len(set(counts.values())) > 1:
                flags.append("%s/%s: groups with unequal hours %r - took max"
                             % (cid, subj, counts))
            if len(by_g) == 1:
                flags.append("%s/%s wk%s: only group %s has this - modelled "
                             "whole-class" % (cid, subj, wk or "-", list(by_g)))
                rows.append((cid, subj, h, tid, blocks_of(list(by_g.values())[0]),
                             1, "", "", wk))
            else:
                rows.append((cid, subj, h, tid,
                             blocks_of(by_g[max(by_g)]), len(by_g), "", "", wk))

    # ---- options (multi-class lessons) ------------------------------------
    # CHOICE options (Spanish/German/Italian/Music/Art) sharing classes are
    # simultaneous bands (H14). Other multi-class lessons - pooled SPORT in
    # the stadium etc. - are joint lessons, each its OWN band (blank cell).
    CHOICE_KEYS = ("إسبان", "ألمان", "المان", "إيطال", "ايطال", "موسيق", "تشكيل")

    def is_choice(subj_id):
        name = subjects.get(subj_id, {}).get("name", "")
        return any(k in name for k in CHOICE_KEYS)

    parent = {}

    def find(a):
        while parent.get(a, a) != a:
            parent[a] = parent.get(parent[a], parent[a])
            a = parent[a]
        return a

    # honest banding: two choice lessons are one band only if last year
    # actually ran them AT THE SAME SLOTS while sharing a class - the data
    # showed pooled options were NOT all simultaneous (different hours,
    # same teacher across pools), so class-sharing alone proves nothing.
    def sig(lid):
        return tuple(sorted((d, p, w) for d, p, wks in opt_cards[lid]
                            for w in wks))

    by_class_sig = {}
    for lid in opt_cards:
        if not is_choice(lessons[lid]["subject"]):
            continue
        parent.setdefault(lid, lid)
        for c in lessons[lid]["classes"]:
            key = (c, sig(lid))
            if key in by_class_sig:
                ra, rb = find(lid), find(by_class_sig[key])
                if ra != rb:
                    parent[ra] = rb
            by_class_sig[key] = lid
    band_of = {}
    n_bands = 0
    for lid in sorted(opt_cards):
        if is_choice(lessons[lid]["subject"]):
            root_ = find(lid)
            if root_ not in band_of:
                n_bands += 1
                band_of[root_] = "OPTB%d" % n_bands
            band_of[lid] = band_of[root_]

    opts = []
    for lid, cards in sorted(opt_cards.items()):
        L = lessons[lid]
        tid = L["teachers"][0] if L["teachers"] else ""
        opts.append(("LY_%s" % lid, L["subject"], tid, len(cards), "",
                     ";".join(L["classes"]), "", band_of.get(lid, "")))

    # ---- option hours = the busier week's card count ----------------------
    opts_fixed = []
    for oid, subj, tid, _h, bl, cls, rt, band in opts:
        lid = oid[3:]
        per_w = {"A": 0, "B": 0}
        for d, p, wks in opt_cards[lid]:
            for w in wks:
                per_w[w] += 1
        opts_fixed.append((oid, subj, tid, max(per_w.values()) or 1, bl, cls,
                           rt, band))
    opts = opts_fixed

    # ---- repair pass: a class booked past the 44-period week --------------
    # Last year sometimes NESTED a short option inside a longer one (pupil
    # subsets we cannot model). Whole-class band booking then overflows the
    # week. Repair: release such a class from its smallest extra band,
    # flagged - the duel loses one binding, not a lesson.
    week_slots = 44
    def class_loads():
        load = collections.defaultdict(lambda: [0, 0])
        for (cid, subj, h, tid, bl, g, rt, core, wk) in rows:
            for i, w in enumerate("AB"):
                if wk in ("", w, "ALT", "ALT2"):
                    load[cid][i] += h
        for oid, subj, tid, h, bl, cls, rt, band in opts:
            for c in cls.split(";"):
                for i in (0, 1):
                    load[c][i] += h
        return load
    for _ in range(10):
        load = class_loads()
        over = [(c, max(e)) for c, e in load.items() if max(e) > week_slots]
        if not over:
            break
        c, _n = over[0]
        mine = [(i, o) for i, o in enumerate(opts) if c in o[5].split(";")]
        if not mine:
            break
        mine.sort(key=lambda io: io[1][3])       # smallest hours first
        i, o = mine[0]
        rest = [x for x in o[5].split(";") if x != c]
        flags.append("class %s over %d slots: released from option %s (%dh)"
                     % (c, week_slots, o[0], o[3]))
        if rest:
            opts[i] = o[:5] + (";".join(rest),) + o[6:]
        else:
            opts.pop(i)

    # ---- teacher contracts: EXACTLY the checker's arithmetic --------------
    # (hours x groups per week from the WRITTEN rows + options both weeks),
    # so H10 is clean by construction.
    tload = collections.defaultdict(lambda: [0.0, 0.0])
    for (cid, subj, h, tid, bl, g, rt, core, wk) in rows:
        if not tid:
            continue
        e = tload[tid]
        th = h * max(1, g)
        if wk in ("", "A"):
            e[0] += th
        if wk in ("", "B"):
            e[1] += th
    for oid, subj, tid, h, bl, cls, rt, band in opts:
        if tid:
            tload[tid][0] += h
            tload[tid][1] += h

    wb = Workbook()

    def sheet(name, header, data):
        ws = wb.create_sheet(name) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = name
        ws.append(header)
        ws.append(["" for _ in header])
        for r in data:
            ws.append(list(r))

    sheet("Teachers",
          ["id", "name", "short", "subjects", "hours", "day_off",
           "training_day", "compact", "travels_with", "notes"],
          [[tid, t["name"], t["short"], "",
            int(math.ceil(sum(tload.get(tid, [0, 0])) / 2.0)) or "",
            "", "", "", "", "contract = last year's average card-hours"]
           for tid, t in sorted(teachers.items())])
    sheet("Classes",
          ["id", "name", "grade", "stream", "is_bac", "cohort", "home_room", "size"],
          [[cid, c["name"],
            next((ch for ch in c["name"] if ch.isdigit()), ""),
            "", "yes" if c["name"].strip().startswith("4") else "",
            "ALL", "", ""]
           for cid, c in sorted(classes.items())])
    # room TYPES and capacities from the real names (Majd 2026-08-25: the
    # first duel put sport at 18:00 because this sheet said every room was
    # an ordinary classroom and every subject had no limits).
    def room_type_of(name):
        n = name.strip()
        for kw, ty in (("ملعب", "gym"), ("فيز", "lab_phys"), ("علوم", "lab_sci"),
                       ("Inf", "it"), ("inf", "it"), ("م ه آلية", "eng_mech"),
                       ("م ه كه", "eng_elec"), ("تقنية", "tech")):
            if kw in n:
                return ty
        return "normal"

    def room_cap(name, ty):
        n = name.strip()
        if ty == "gym":
            return 99
        if n in ("فيز2", "فيز3"):
            return 40
        if ty in ("lab_sci",):
            return 30
        if ty in ("lab_phys", "it", "tech", "eng_mech", "eng_elec"):
            return 20
        if n.lower() == "group":
            return 20
        return 40

    sheet("Rooms",
          ["id", "name", "type", "capacity", "zone", "notes"],
          [[rid, r["name"], room_type_of(r["name"]),
            room_cap(r["name"], room_type_of(r["name"])), "",
            "type+capacity from the room name (Majd's house rules)"]
           for rid, r in sorted(rooms.items())])
    # subject ATTRIBUTES from the name: the ministry limits Majd's real
    # sheet already carries (daylight for sport, labs, maths before 16h...)
    def attrs_of(name):
        n = name.strip()
        def has(*kw):
            return any(k in n for k in kw)
        if has("بدنية", "رياضة") and not has("رياضيات"):
            # H15 daylight: no stadium lighting after 16:00 = period 8
            return ("easy", "gym", 8, "", "yes", "yes", "", "")
        if has("فيزيائية"):
            # theory sits in an ordinary classroom; only the GROUP (TP)
            # rows get the lab - written per row below (Majd's P9 split)
            return ("hard", "normal", "", 9, "", "", "", "scientific")
        if has("الحياة", "حياة"):
            return ("medium", "normal", "", "", "", "", "", "scientific")
        if has("اعلامية", "إعلامية", "خوارزميات", "المعلومات", "قواعد", "الشبكات"):
            return ("medium", "it", "", "", "", "", "", "scientific")
        if has("هنسة آلية", "هندسة آلية"):
            return ("medium", "eng_mech", "", "", "", "", "", "scientific")
        if has("هنسة كهربائية", "هندسة كهربائية"):
            return ("medium", "eng_elec", "", "", "", "", "", "scientific")
        if has("تقنية", "تكنولوجية"):
            return ("medium", "tech", "", "", "", "", "", "scientific")
        if has("رياضيات"):
            return ("hard", "normal", "", 8, "", "", "", "scientific")
        if has("فلسفة"):
            return ("hard", "normal", "", 9, "", "", "", "literary")
        if has("عربية", "فرنسية", "نقليزية", "إسبان", "ألمان", "يطال"):
            return ("medium", "normal", "", "", "", "", "", "literary")
        if has("تاريخ", "جغراف", "مدنية", "تفكير", "إسلام", "اقتصاد", "تصرف"):
            return ("medium", "normal", "", "", "", "", "", "social")
        if has("موسيق", "تشكيل"):
            return ("easy", "normal", "", "", "yes", "", "", "")
        return ("medium", "normal", "", "", "", "", "", "")

    subj_rows = []
    for sid, sb in sorted(subjects.items()):
        diff, rt, latest, avoid, exempt, g24, notaft, nature = attrs_of(sb["name"])
        subj_rows.append([sid, sb["name"], sb["short"], diff, rt, latest,
                          avoid, exempt, g24, notaft, nature])
    sheet("Subjects",
          ["id", "name", "short", "difficulty", "room_type", "latest_period",
           "avoid_after", "minmax_exempt", "gap24", "not_after", "nature"],
          subj_rows)
    lab_row = {}
    for sid, sb in subjects.items():
        n = sb["name"]
        if "فيزيائية" in n:
            lab_row[sid] = "lab_phys"
        elif "الحياة" in n or "حياة" in n:
            lab_row[sid] = "lab_sci"
    rows = [list(r) for r in rows]
    for r in rows:
        if r[5] and int(r[5]) > 1 and r[1] in lab_row:
            r[6] = lab_row[r[1]]          # TP group rows -> the lab
    sheet("Curriculum",
          ["class_id", "subject_id", "hours", "teacher_id", "blocks", "groups",
           "room_type", "core", "week"],
          rows)
    sheet("Options",
          ["id", "subject_id", "teacher_id", "hours", "blocks", "classes",
           "room_type", "band"],
          opts)
    wb.save(OUT)
    print("wrote %s" % OUT)
    print("  %d teachers, %d classes, %d rooms, %d subjects"
          % (len(teachers), len(classes), len(rooms), len(subjects)))
    print("  %d curriculum rows (%d with groups>1, %d fortnight A/B), "
          "%d option groups, %d cards skipped"
          % (len(rows), sum(1 for r in rows if r[5] > 1),
             sum(1 for r in rows if r[8]), len(opts), skipped))
    for f in flags[:10]:
        print("  FLAG:", f)


if __name__ == "__main__":
    main()
