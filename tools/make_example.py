# -*- coding: utf-8 -*-
"""Build examples/school_example.xlsx - a SMALL FICTIONAL school showing
every feature of the data format, for people finding this project online.

All names are invented. No real person or school appears here - the real
data lives in data/ and never leaves the machine (see docs/PRIVACY.md).

The example must always PASS the checker:
    python solver/data.py examples/school_example.xlsx

    python tools/make_example.py
"""
import os
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "examples", "school_example.xlsx")

INTRO = [
    ("SCHOOL TIMETABLE SOLVER - EXAMPLE DATA FILE", True),
    ("", False),
    ("A small FICTIONAL school showing how to fill the real workbook.", False),
    ("Built primarily for TUNISIAN secondary schools (ministry circular", False),
    ("51/2018 rules) exporting to aSc TimeTables 2013 via XML import.", False),
    ("", False),
    ("Row 2 of every sheet is a grey HINT row - the program skips it.", False),
    ("Start your own data on row 3.", False),
    ("", False),
    ("Features demonstrated:", False),
    ("  - blocks '2+1+1': a double + two singles, each on its own day,", False),
    ("    never straddling the lunch break (H9)", False),
    ("  - groups=2: the class splits in half; hours are PER GROUP and", False),
    ("    the teacher teaches both halves (T43)", False),
    ("  - week=A / B: fortnightly lessons - week A / week B (T42)", False),
    ("  - week=ALT: the two groups take turns - group 1 week A,", False),
    ("    group 2 week B (the Tunisian TP fortnight practice)", False),
    ("  - Options sheet: pooled option groups (Spanish/German...) from", False),
    ("    several same-year classes, running simultaneously (H14)", False),
    ("  - day_off: blank = the solver picks the free day; a written", False),
    ("    day = fixed; (none) = no day off (H7)", False),
    ("  - Locked: pin lessons; the solver builds around them", False),
    ("  - Weights sheet (auto-added on first run): every comfort rule's", False),
    ("    priority, editable in Excel; write HARD to make it unbreakable", False),
    ("", False),
    ("Check it: python solver/data.py  examples/school_example.xlsx", False),
    ("Solve it: python solver/solve.py examples/school_example.xlsx", False),
]

# name -> (cols, hint, rows). A trailing "notes" column is free text.
SHEETS = {
    "Teachers": (
        ["id", "name", "short", "subjects", "hours", "day_off",
         "training_day", "compact", "travels_with", "notes"],
        ["never changes", "any language", "for the grid", "info only",
         "contract hours/week", "blank = solver picks; day = fixed; (none)",
         "weekly training day, kept free", "yes = few full days",
         "teacher id sharing transport", "free text"],
        [
            ["T1", "Ali Example", "ALI", "", 18, "", "", "", "", "blank day_off: the solver picks the day"],
            ["T2", "Sara Sample", "SAR", "", 18, "Wed", "", "", "", "written day = FIXED day off"],
            ["T3", "Omar Demo", "OMA", "", 18, "", "Thu", "", "", "training Thu - day off never adjacent (H18)"],
            ["T4", "Lina Fictive", "LIN", "", 18, "(none)", "", "yes", "", "compact: prefers few full days"],
            ["T5", "Nora Invent", "NOR", "", 18, "", "", "", "", "teaches the Spanish option"],
            ["T6", "Karim Nobody", "KAR", "", 18, "", "", "", "T5", "shared transport with T5 (S21)"],
        ]),
    "Classes": (
        ["id", "name", "grade", "stream", "is_bac", "cohort", "home_room", "size"],
        ["never changes", "printed name", "1-4", "stream code",
         "yes = final year (protected afternoons)", "ALL", "preferred room",
         "pupil count - drives H16 split warnings"],
        [
            ["C1", "1S1", 1, "COMMON", "", "ALL", "R1", 32],
            ["C2", "1S2", 1, "COMMON", "", "ALL", "R2", 22],
            ["C3", "4M1", 4, "MATHS", "yes", "ALL", "", 28],
        ]),
    "Rooms": (
        ["id", "name", "type", "capacity", "zone", "notes"],
        ["never changes", "printed name", "normal/lab_sci/gym/it/...",
         "seats", "rooms sharing a zone are close (T45)", "free text"],
        [
            ["R1", "Salle 1", "normal", 40, "A", ""],
            ["R2", "Salle 2", "normal", 40, "A", ""],
            ["R3", "Salle 3", "normal", 40, "B", "walks between zones are scored"],
            ["L1", "Labo SVT", "lab_sci", 20, "B", ""],
            ["G1", "Stade", "gym", 99, "B", ""],
        ]),
    "Subjects": (
        ["id", "name", "short", "difficulty", "room_type", "latest_period",
         "avoid_after", "minmax_exempt", "gap24", "not_after", "nature", "notes"],
        ["never changes", "printed name", "for the grid", "hard/medium/easy",
         "default room kind", "hard daylight limit (H15)",
         "soft: prefer not after this period", "yes = exempt from min-2h",
         "yes = 24h between sessions", "ids this must not follow",
         "literary/scientific/social", "free text"],
        [
            ["MATH", "Mathématiques", "Math", "hard", "normal", "", 8, "", "", "", "scientific", "soft: not after 16:00"],
            ["ARAB", "Arabe", "Ar", "medium", "normal", "", "", "", "", "", "literary", ""],
            ["HIST", "Histoire", "Hist", "medium", "normal", "", "", "", "", "", "social", ""],
            ["SVT", "Sciences SVT", "SVT", "medium", "normal", "", "", "", "", "", "scientific", ""],
            ["SVT_TP", "TP SVT", "TP", "medium", "lab_sci", "", "", "", "", "", "scientific", "_TP suffix auto-links: never same day as SVT"],
            ["SPORT", "Sport", "EPS", "easy", "gym", 8, "", "yes", "yes", "", "", "daylight only + 24h between sessions"],
            ["ESP", "Espagnol", "Esp", "medium", "normal", "", "", "", "", "", "", "option subject - see Options sheet"],
            ["ALL", "Allemand", "All", "medium", "normal", "", "", "", "", "", "", "option subject - see Options sheet"],
        ]),
    "Curriculum": (
        ["class_id", "subject_id", "hours", "teacher_id", "blocks", "groups",
         "room_type", "core", "week", "notes"],
        ["from Classes", "from Subjects", "per week - PER GROUP if groups>1",
         "from Teachers (blank = no teacher rules)", "e.g. 2+1+1",
         "1 = whole class, 2 = split halves", "blank = subject default",
         "yes = 3/4 of hours in the morning", "blank / A / B / ALT",
         "free text"],
        [
            ["C1", "MATH", 4, "T1", "2+1+1", 1, "", "yes", "", ""],
            ["C1", "ARAB", 3, "T2", "2+1", 1, "", "", "", ""],
            ["C1", "SVT", 2, "T3", "2", 1, "", "", "", "theory in a normal room"],
            ["C1", "SVT_TP", 1, "T3", "1", 2, "", "", "ALT", "split TP: group 1 week A, group 2 week B"],
            ["C1", "SPORT", 2, "T4", "1+1", 1, "", "", "", ""],
            ["C2", "MATH", 4, "T1", "2+1+1", 1, "", "yes", "", ""],
            ["C2", "ARAB", 3, "T2", "2+1", 1, "", "", "", ""],
            ["C2", "SPORT", 2, "T4", "1+1", 1, "", "", "", ""],
            ["C3", "MATH", 5, "T1", "2+2+1", 1, "", "yes", "", "bac class: afternoons protected (S17)"],
            ["C3", "HIST", 2, "T3", "1+1", 1, "", "", "", ""],
            ["C3", "HIST", 1, "T3", "1", 1, "", "", "A", "fortnightly extra hour, week A only"],
        ]),
    "Options": (
        ["id", "subject_id", "teacher_id", "hours", "blocks", "classes", "room_type"],
        ["one row per option GROUP", "ESP/ALL/ITA/MUS/TASH", "from Teachers",
         "hours per week", "e.g. 2 or 1+1", "same-year classes, C1;C2",
         "blank = subject default"],
        [
            ["OPT1_ESP", "ESP", "T5", 2, "", "C1;C2", ""],
            ["OPT1_ALL", "ALL", "T6", 2, "", "C1;C2", ""],
        ]),
    "Unavailable": (
        ["teacher_id", "day", "period", "hard", "reason"],
        ["from Teachers", "or * for every day", "or * for whole day",
         "yes = never; no = prefer not", "printed in the report"],
        [
            ["T2", "Mon", 1, "yes", "example: school council meeting"],
        ]),
    "Locked": (
        ["class_id", "subject_id", "day", "period", "room_id", "why"],
        ["from Classes", "from Subjects", "day", "period number",
         "room id or blank", "free text"],
        [
            ["C1", "SPORT", "Tue", 3, "", "example: stadium shared with another school"],
        ]),
}


def main():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "READ ME"
    ws.column_dimensions["A"].width = 78
    for i, (text, big) in enumerate(INTRO, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=True, size=14, color="1F4E78") if big else Font(size=11)

    grey = Font(italic=True, color="808080")
    head = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEBF7")
    for name, (cols, hint, rows) in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(cols)
        ws.append(hint)
        for c in ws[1]:
            c.font = head
            c.fill = fill
        for c in ws[2]:
            c.font = grey
        for r in rows:
            ws.append(r)
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions[chr(ord("A") + len(cols) - 1)].width = 40
    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
