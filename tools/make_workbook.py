"""Generate data/school.xlsx - the friendly input file the user actually edits.

CSV stays the storage format under the hood, but nobody has to look at it.
This workbook has one sheet per table, coloured headers, frozen panes,
dropdown menus on every field with fixed choices, and a HOW TO sheet.
Re-run this ONLY to create a blank workbook. It will refuse to overwrite
a workbook that already has data in it.
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "data", "school.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
NOTE_FONT = Font(color="808080", italic=True, size=10)

# sheet -> (columns, human hints, dropdowns {col: choices}, width overrides)
SHEETS = {
    "Teachers": dict(
        cols=["id", "name", "short", "subjects", "hours", "day_off",
              "training_day", "compact", "notes"],
        hint=["T01, T02...", "full name as printed", "2-4 letters for the grid",
              "codes separated by ;", "contracted hours/week",
              "pick from list", "pedagogical training day - must stay empty",
              "yes = pack into fewer days (long journey). blank = ministry default",
              "free text - solver ignores"],
        dv={"day_off": "Mon,Tue,Wed,Thu,Fri,Sat,(none)",
            "training_day": "Mon,Tue,Wed,Thu,Fri,Sat,(none)",
            "compact": "yes,no"},
        width={"name": 28, "subjects": 24, "notes": 34}),
    "Classes": dict(
        cols=["id", "name", "grade", "stream", "size", "is_bac",
              "home_room", "cohort"],
        hint=["C01, C02...", "as printed e.g. 4رياضيات1", "level 1-4",
              "which stream - drives the curriculum table",
              "number of pupils - H16: no split at 24 or fewer",
              "yes for 4th year - bac has its own ministry rules",
              "room id or blank", "AM/PM/ALL"],
        dv={"cohort": "AM,PM,ALL", "grade": "1,2,3,4", "is_bac": "yes,no",
            "stream": "COMMON,LETTERS,SCIENCES,MATHS,EXPSCI,TECHSCI,CS,ECONOMY,IT_TECH"},
        width={"name": 26}),
    "Rooms": dict(
        cols=["id", "name", "type", "capacity", "zone", "computers", "notes"],
        hint=["R01, R02...", "as printed", "pick from list", "max pupils",
              "where it is, e.g. A-rez / A-1 / stade",
              "IT labs only - pupils must not exceed twice this",
              "free text, e.g. shared on Tuesdays"],
        dv={"type": "normal,lab_phys,lab_chem,lab_sci,it,gym,tech,music,arts"},
        width={"name": 26, "zone": 14, "notes": 34}),
    "Zones": dict(
        cols=["from", "to", "walk_minutes"],
        hint=["a zone from the Rooms sheet", "another zone",
              "minutes to walk between them"],
        dv={},
        width={"walk_minutes": 14}),
    "Subjects": dict(
        cols=["id", "name", "short", "difficulty", "room_type", "latest_period",
              "avoid_after", "minmax_exempt"],
        hint=["MATH, PHYS...", "Arabic or French", "grid abbreviation",
              "drives morning rule", "room needed",
              "last period allowed, blank = any (Sport=8, ends 16:00)",
              "SOFT: prefer not after this period (Maths=8, Physics=9)",
              "yes = PE/optional: exempt from the min-2h rules (circular I.2)"],
        dv={"difficulty": "hard,medium,easy",
            "room_type": "normal,lab_phys,lab_chem,lab_sci,it,gym,tech",
            "minmax_exempt": "yes,no"},
        width={"name": 30, "avoid_after": 13, "minmax_exempt": 15}),
    "Curriculum": dict(
        cols=["class_id", "subject_id", "hours", "teacher_id", "blocks",
              "groups", "room_type"],
        hint=["from Classes", "from Subjects", "periods per week",
              "from Teachers (blank = solver picks)",
              "e.g. 2+2+1 or 1+1+1",
              "1 = whole class, 2 = split. Ministry: do not split at 24 pupils or fewer",
              "blank = subject default"],
        dv={"room_type": "normal,lab_phys,lab_chem,lab_sci,it,gym,tech",
            "groups": "1,2,3"},
        width={"blocks": 16}),
    "Unavailable": dict(
        cols=["teacher_id", "day", "period", "hard", "reason"],
        hint=["from Teachers", "or * for every day", "or * for whole day",
              "yes = never. no = prefer not",
              "printed in the report so you can defend it"],
        dv={"day": "Mon,Tue,Wed,Thu,Fri,Sat,*", "hard": "yes,no"},
        width={"reason": 40}),
    "Locked": dict(
        cols=["class_id", "subject_id", "day", "period", "room_id", "why"],
        hint=["from Classes", "from Subjects", "pick from list", "period number",
              "room id or blank", "free text - why you pinned it"],
        dv={"day": "Mon,Tue,Wed,Thu,Fri,Sat"},
        width={"why": 40}),
}

HOWTO = [
    ("HOW TO FILL THIS FILE", True),
    ("", False),
    ("One sheet per kind of thing. Fill them in this order:", False),
    ("   1. Teachers   2. Classes   3. Rooms   4. Subjects", False),
    ("   5. Curriculum (the big one)   6. Unavailable (only if needed)", False),
    ("", False),
    ("Row 2 of every sheet is a grey HINT row explaining each column.", False),
    ("Leave it there. The program skips it. Start typing on row 3.", False),
    ("", False),
    ("Blue columns have a dropdown - click the cell and pick. Do not type", False),
    ("your own value there or the checker will complain.", False),
    ("", False),
    ("IDs must be short and never change. If a teacher leaves, delete the", False),
    ("row - do not renumber anything. That is what makes re-running safe.", False),
    ("", False),
    ("Arabic and French names are both fine anywhere in the name columns.", False),
    ("", False),
    ("The LOCKED sheet is how you steer the solver. Leave it empty at first.", False),
    ("After a run, pin anything you like there and re-run - the solver keeps", False),
    ("your pinned lessons exactly and rebuilds everything else around them.", False),
    ("", False),
    ("WHEN YOU ARE DONE: save, then double-click  check_data.bat", False),
    ("It tells you in plain language what is missing or wrong. It never", False),
    ("changes your file.", False),
]


def build():
    if os.path.exists(OUT):
        from openpyxl import load_workbook
        wb_old = load_workbook(OUT)
        for name in SHEETS:
            if name in wb_old.sheetnames and wb_old[name].max_row > 2:
                sys.exit(f"REFUSING: {OUT}\n  sheet '{name}' already has data. "
                         f"Delete or rename the file first if you really want a blank one.")

    wb = Workbook()
    ws = wb.active
    ws.title = "HOW TO"
    ws.column_dimensions["A"].width = 78
    for i, (text, big) in enumerate(HOWTO, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=True, size=14, color="1F4E78") if big else Font(size=11)

    for name, spec in SHEETS.items():
        ws = wb.create_sheet(name)
        cols, hints = spec["cols"], spec["hint"]
        for j, (col, hint) in enumerate(zip(cols, hints), start=1):
            h = ws.cell(row=1, column=j, value=col)
            h.fill, h.font = HEAD_FILL, HEAD_FONT
            h.alignment = Alignment(horizontal="center")
            n = ws.cell(row=2, column=j, value=hint)
            n.font = NOTE_FONT
            letter = get_column_letter(j)
            ws.column_dimensions[letter].width = spec["width"].get(col, max(12, len(col) + 4))
        for col, choices in spec["dv"].items():
            j = cols.index(col) + 1
            letter = get_column_letter(j)
            dv = DataValidation(type="list", formula1=f'"{choices}"', allow_blank=True)
            dv.error = f"Pick one of: {choices}"
            dv.errorTitle = "Not a valid value"
            ws.add_data_validation(dv)
            dv.add(f"{letter}3:{letter}600")
            ws.cell(row=1, column=j).fill = PatternFill("solid", fgColor="2E75B6")
        ws.freeze_panes = "A3"

    wb.save(OUT)
    print(f"created {OUT}")
    print(f"sheets: {', '.join(['HOW TO'] + list(SHEETS))}")


if __name__ == "__main__":
    build()
