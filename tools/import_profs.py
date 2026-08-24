"""Read the official ministry teacher list and fill the Teachers sheet.

The file is an administrative document (Regional Delegation letterhead, a
title block, then a table), not a clean spreadsheet - so the header row is
found by looking for its Arabic column names rather than assumed to be row 1.

    python tools/import_profs.py [data/profs.xlsx]

DATA MINIMISATION - deliberate, not an oversight:

The official list carries a column "المعرف الوحيد" - the teacher's national
identity number. **We do not copy it.** A timetable does not need government ID
numbers, and data you never store is data that can never leak. Only these are
carried across:

    name       - so the printed timetable is readable
    subject    - so we know who can teach what
    id         - a code WE generate (T001...), stable across re-runs

Recruitment dates, institution codes and ID numbers are all left behind.
"""
import os
import sys

# Windows consoles default to cp1252, which cannot encode Arabic and raises
# UnicodeEncodeError mid-print. Force UTF-8 so real names are printable.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_SRC = os.path.join(HERE, "data", "profs.xlsx")
TARGET = os.path.join(HERE, "data", "school.xlsx")

# Arabic column headings in the ministry format -> what we call them
COLUMNS = {
    "اسم الاستاذ ولقبه": "name",
    "مادة التدريس": "subject",
    "ملاحظات": "notes",
    "ع/ر": "serial",
}
# Columns deliberately NOT imported, and why. Printed at the end so the
# omission is visible rather than silent.
DROPPED = {
    "المعرف الوحيد": "national identity number - not needed for a timetable",
    "رمز المؤسسة التربوية": "institution code - same for everyone",
    "إسم المؤسسة التربوية": "institution name - same for everyone",
    "تاريخ الانتاداب": "recruitment date - not needed",
    "تاريخ اللالتحاق بالمؤسسة": "date joined - not needed",
}

# Arabic subject name -> our subject code. Extend as new ones appear.
SUBJECTS = {
    "عربية": ("ARAB", "Arabe"),
    "فرنسية": ("FREN", "Francais"),
    "انقليزية": ("ENGL", "Anglais"),
    "إنقليزية": ("ENGL", "Anglais"),
    "رياضيات": ("MATH", "Mathematiques"),
    "علوم فيزيائية": ("PHYS", "Sciences Physiques"),
    "علوم الحياة والأرض": ("SVT", "Sciences Naturelles"),
    "تاريخ وجغرافيا": ("HIST", "Histoire-Geo"),
    "تربية إسلامية": ("ISLA", "Education Islamique"),
    "تربية بدنية": ("SPORT", "Education Physique"),
    "إعلامية": ("IT", "Informatique"),
    "تفكير إسلامي": ("PISL", "Pensee Islamique"),
    "فلسفة": ("PHIL", "Philosophie"),
    "اسبانية": ("ESP", "Espagnol"),
    "ألمانية": ("ALL", "Allemand"),
    "إيطالية": ("ITA", "Italien"),
    "تربية تشكيلية": ("TASH", "Education Plastique"),
    "تربية موسيقية": ("MUS", "Education Musicale"),
    "تكنولوجيا": ("TECH", "Technologie"),
    "اقتصاد وتصرف": ("ECO", "Economie et Gestion"),
    # --- abbreviated / spelling variants found in the real 2026 list ---
    "ع.فيزيائية": ("PHYS", "Sciences Physiques"),
    "أنقليزية": ("ENGL", "Anglais"),
    "ايطالية": ("ITA", "Italien"),
    "علوم ح. أ.": ("SVT", "Sciences Naturelles"),
    "اقتصاد": ("ECO", "Economie"),
    "تصرف": ("GEST", "Gestion"),
    "تربية مدنية": ("CIV", "Education Civique"),
}


def find_header(rows):
    """Return (index, {col_number: our_name}) for the real table header."""
    for i, r in enumerate(rows[:40]):
        cells = ["" if v is None else str(v).strip() for v in r]
        mapping = {}
        for j, c in enumerate(cells):
            if c in COLUMNS:
                mapping[j] = COLUMNS[c]
        if "name" in mapping.values() and "subject" in mapping.values():
            return i, mapping
    return None, {}


def main():
    from openpyxl import load_workbook

    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    if not os.path.exists(src):
        sys.exit("No file at " + src)

    ws = load_workbook(src, read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr, cols = find_header(rows)
    if hdr is None:
        sys.exit("Could not find the table header. Expected a row containing "
                 "both 'اسم الاستاذ ولقبه' and 'مادة التدريس'.")
    print("Header found on row %d." % (hdr + 1))

    teachers, blank, unknown = [], 0, {}
    for r in rows[hdr + 1:]:
        rec = {}
        for j, key in cols.items():
            v = r[j] if j < len(r) else None
            rec[key] = "" if v is None else str(v).strip()
        if not rec.get("name"):
            if any(str(v).strip() for v in r if v is not None):
                blank += 1          # a numbered row with no teacher yet
            continue
        ar = rec.get("subject", "")
        code, _fr = SUBJECTS.get(ar, ("", ""))
        if ar and not code:
            unknown[ar] = unknown.get(ar, 0) + 1
        teachers.append(dict(name=rec["name"], subject=code or "",
                             arabic_subject=ar, notes=rec.get("notes", "")))

    print("Teachers with a name: %d" % len(teachers))
    if blank:
        print("Numbered rows with no teacher yet: %d  (vacancies / not arrived)" % blank)

    # stable ids, assigned in file order
    for n, t in enumerate(teachers, start=1):
        t["id"] = "T%03d" % n

    if unknown:
        print("")
        print("UNKNOWN SUBJECTS - add these to SUBJECTS in this script:")
        for ar, n in sorted(unknown.items(), key=lambda kv: -kv[1]):
            print("   %-30s x%d" % (ar, n))
        print("")
        print("Nothing written. Fix the mapping first, then re-run.")
        return 1

    if not os.path.exists(TARGET):
        sys.exit("No data/school.xlsx yet - run tools/make_workbook.py first.")
    wb = load_workbook(TARGET)
    sheet = wb["Teachers"]
    if sheet.max_row > 2:
        print("")
        print("Teachers sheet already has %d rows." % (sheet.max_row - 2))
        print("Refusing to overwrite. Clear it first if you want a fresh import.")
        return 1
    for t in teachers:
        # hours and day_off deliberately left blank - they are not in the
        # ministry list and must not be invented.
        sheet.append([t["id"], t["name"], "", t["subject"], "", "", t["notes"]])
    if "_DEMO" in wb.sheetnames:
        del wb["_DEMO"]          # this is real data now, not the demo
    wb.save(TARGET)

    print("")
    print("Wrote %d teachers into data/school.xlsx (Teachers sheet)." % len(teachers))
    print("")
    print("NOT imported, on purpose:")
    for ar, why in DROPPED.items():
        print("   %-28s %s" % (ar, why))
    print("")
    print("STILL BLANK - these are not in the ministry list and were not invented:")
    print("   hours     contracted hours per week")
    print("   day_off   which day each teacher is off")
    print("   short     grid abbreviation (can be generated later)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
