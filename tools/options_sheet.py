# -*- coding: utf-8 -*-
"""Add the H14 'Options' sheet to data/school.xlsx (empty, for Majd).

One row per option GROUP: the pupils who chose that option, pooled from the
listed same-year classes. Groups that share a class run SIMULTANEOUSLY (a
band) - Majd's rule: while options run, a pupil can only be in another
option. Refuses to overwrite a sheet that already has data.

    python tools/options_sheet.py
"""
import os
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(HERE, "data", "school.xlsx")


def main():
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(XLSX)
    if "Options" in wb.sheetnames:
        ws = wb["Options"]
        if ws.max_row > 2:
            sys.exit("REFUSING: the Options sheet already has data.")
        wb.remove(ws)
    ws = wb.create_sheet("Options")
    ws.append(["id", "subject_id", "teacher_id", "hours", "blocks",
               "classes", "room_type"])
    ws.append(["e.g. OPT3_ESP", "ESP / ALL / ITA / MUS / TASH",
               "from Teachers", "hours per week", "e.g. 2 or 1+1",
               "same-year classes, like C13;C14;C15",
               "blank = subject default"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
    for c in ws[2]:
        c.font = Font(italic=True, color="808080")
    ws.column_dimensions["F"].width = 28
    wb.save(XLSX)
    print("Options sheet added (empty). Groups sharing a class run "
          "simultaneously - fill it whenever ready.")


if __name__ == "__main__":
    main()
