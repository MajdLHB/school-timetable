"""Pre-fill the Rooms sheet from last year's aSc export.

Majd (2026-08-24): "rooms usually same as last year, fill data in, will see
them and verify". So: the 45 room names come from the real export, the TYPE
of each room is inferred from its name, and every row is marked PREFILL in
the notes so nobody mistakes a guess for a fact. Majd edits the sheet in
Excel and simply deletes the note once a row is checked.

Refuses to run if the Rooms sheet already has data.

    python tools/prefill_rooms.py
"""
import os
import sys
import xml.etree.ElementTree as ET

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(HERE, "data", "school.xlsx")
LAST = os.path.join(HERE, "data", "reference", "last-year-2025-26", "last-yeah.xml")

# name prefix -> (room type, human reason)
# aSc labels the export windows-1252 but writes Arabic as cp1256 (see T3).
RULES = [
    ("علوم", "lab_sci", "SVT lab"),
    ("فيز", "lab_phys", "physics lab"),
    ("inf", "it", "IT room"),
    ("تقنية", "tech", "technology room"),
    ("م ه", "tech", "engineering room (mech/elec)"),
    ("ملعب", "gym", "stadium subdivision"),
    ("ق", "normal", "ordinary classroom"),
]


def room_type(name):
    low = name.strip().lower()
    for prefix, rtype, why in RULES:
        if low.startswith(prefix):
            return rtype, why
    return "normal", "no name match - GUESSED normal"


def main():
    from openpyxl import load_workbook

    raw = open(LAST, "rb").read().decode("cp1256", errors="replace")
    rooms = ET.fromstring(raw).findall(".//classroom")
    if not rooms:
        sys.exit("No classrooms found in " + LAST)

    wb = load_workbook(XLSX)
    ws = wb["Rooms"]
    if ws.max_row > 2:
        sys.exit("REFUSING: the Rooms sheet already has data. Edit it in "
                 "Excel instead, or clear it first if you want a re-fill.")

    counts = {}
    for i, r in enumerate(rooms, 1):
        name = (r.get("name") or "").strip()
        short = (r.get("short") or name).strip()
        rtype, why = room_type(name)
        counts[rtype] = counts.get(rtype, 0) + 1
        note = "PREFILL from last year - type '%s' guessed (%s). Check and delete this note." % (rtype, why)
        if name == "Group":
            note = "PREFILL - last year had a room literally named 'Group'. Probably a dummy for group lessons. Confirm or DELETE this row."
        # Columns: id, name, type, capacity, zone, computers, notes
        ws.append(["R%02d" % i, name, rtype, 35, "", "", note])

    wb.save(XLSX)
    print("Wrote %d rooms into the Rooms sheet of %s" % (len(rooms), XLSX))
    for t in sorted(counts):
        print("  %-9s %d" % (t, counts[t]))
    print("Every row carries a PREFILL note. Majd: check the types, fill the")
    print("computers column for IT rooms (ministry: pupils <= 2x computers),")
    print("and delete each note once the row is verified.")


if __name__ == "__main__":
    main()
